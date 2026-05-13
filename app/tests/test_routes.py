"""Tests para las rutas de la aplicación Almacén."""

import io
import os
import tempfile
import pytest
from app import create_app, db
from app.models import User, Producto, Entrada, Salida, OrdenCompra


@pytest.fixture(scope="function")
def app():
    """Create app with temporary SQLite file for testing."""
    tmp_db = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
    tmp_db.close()
    os.environ["DATABASE_URL"] = f"sqlite:///{tmp_db.name}"

    application = create_app(testing=True)
    application.config["WTF_CSRF_ENABLED"] = False

    with application.app_context():
        db.create_all()
        admin = User(username="cponce123.com@gmail.com")
        admin.set_password("Hadrones456%")
        db.session.add(admin)
        db.session.commit()
        yield application
        db.drop_all()

    os.unlink(tmp_db.name)
    # Clean up WAL/SHM files
    for ext in ("-wal", "-shm"):
        p = tmp_db.name + ext
        if os.path.exists(p):
            os.unlink(p)
    del os.environ["DATABASE_URL"]


@pytest.fixture(scope="function")
def client(app):
    """Test client."""
    return app.test_client()


@pytest.fixture(scope="function")
def auth_client(client):
    """Authenticated test client."""
    client.post("/login", data={"username": "cponce123.com@gmail.com", "password": "Hadrones456%"})
    return client


@pytest.fixture(scope="function")
def sample_productos(app):
    """Create sample products for testing."""
    with app.app_context():
        p1 = Producto(codigo="P001", descripcion="Tornillo M8x30", um="UND",
                      familia="FERRETERIA", stock_actual=100, stock_minimo=10)
        p2 = Producto(codigo="P002", descripcion="Tuerca M8", um="UND",
                      familia="FERRETERIA", stock_actual=50, stock_minimo=5)
        p3 = Producto(codigo="P003", descripcion="Arandela M8", um="UND",
                      familia="FERRETERIA", stock_actual=0, stock_minimo=10)
        db.session.add_all([p1, p2, p3])
        db.session.commit()


# ===========================================================================
# Tests de autenticación
# ===========================================================================

class TestAuth:
    def test_login_get(self, client):
        resp = client.get("/login")
        assert resp.status_code == 200

    def test_login_post_success(self, client):
        resp = client.post("/login", data={"username": "cponce123.com@gmail.com", "password": "Hadrones456%"},
                           follow_redirects=True)
        assert resp.status_code == 200
        assert b"Inicio de sesi" in resp.data

    def test_login_post_fail(self, client):
        resp = client.post("/login", data={"username": "cponce123.com@gmail.com", "password": "wrong"},
                           follow_redirects=True)
        assert resp.status_code == 200
        assert b"incorrectos" in resp.data

    def test_logout(self, auth_client):
        resp = auth_client.get("/logout", follow_redirects=True)
        assert resp.status_code == 200
        assert b"iniciar sesi" in resp.data.lower()

    def test_protected_redirects_to_login(self, client):
        resp = client.get("/productos", follow_redirects=False)
        assert resp.status_code == 302
        assert "/login" in resp.location

    def test_cambiar_contrasena_get(self, auth_client):
        resp = auth_client.get("/cambiar-contrasena")
        assert resp.status_code == 200
        assert b"Cambiar mi contrase" in resp.data

    def test_cambiar_contrasena_post_success(self, auth_client):
        """El usuario cambia su contraseña correctamente."""
        resp = auth_client.post("/cambiar-contrasena", data={
            "current_password": "Hadrones456%",
            "new_password": "NuevaPass123",
            "new_password2": "NuevaPass123",
        }, follow_redirects=True)
        assert resp.status_code == 200
        assert b"cambiada correctamente" in resp.data

        # Verificar que puede iniciar sesión con la nueva contraseña
        auth_client.get("/logout", follow_redirects=True)
        resp = auth_client.post("/login", data={
            "username": "cponce123.com@gmail.com",
            "password": "NuevaPass123",
        }, follow_redirects=True)
        assert b"Inicio de sesi" in resp.data

    def test_cambiar_contrasena_wrong_current(self, auth_client):
        """Error si la contraseña actual es incorrecta."""
        resp = auth_client.post("/cambiar-contrasena", data={
            "current_password": "wrongpass",
            "new_password": "NuevaPass123",
            "new_password2": "NuevaPass123",
        }, follow_redirects=True)
        assert resp.status_code == 200
        assert b"actual es incorrecta" in resp.data

    def test_cambiar_contrasena_mismatch(self, auth_client):
        """Error si las nuevas contraseñas no coinciden."""
        resp = auth_client.post("/cambiar-contrasena", data={
            "current_password": "Hadrones456%",
            "new_password": "NuevaPass123",
            "new_password2": "OtraPass456",
        }, follow_redirects=True)
        assert resp.status_code == 200
        assert b"no coinciden" in resp.data

    def test_cambiar_contrasena_short_password(self, auth_client):
        """Error si la nueva contraseña es demasiado corta."""
        resp = auth_client.post("/cambiar-contrasena", data={
            "current_password": "Hadrones456%",
            "new_password": "ab",
            "new_password2": "ab",
        }, follow_redirects=True)
        assert resp.status_code == 200
        assert b"al menos 4 caracteres" in resp.data

    def test_cambiar_contrasena_requires_login(self, client):
        """Redirige al login si no está autenticado."""
        resp = client.get("/cambiar-contrasena", follow_redirects=False)
        assert resp.status_code == 302
        assert "/login" in resp.location


# ===========================================================================
# Tests de Dashboard
# ===========================================================================

class TestDashboard:
    def test_dashboard(self, auth_client, sample_productos):
        resp = auth_client.get("/")
        assert resp.status_code == 200


# ===========================================================================
# Tests de Productos CRUD
# ===========================================================================

class TestProductos:
    def test_list_productos(self, auth_client, sample_productos):
        resp = auth_client.get("/productos")
        assert resp.status_code == 200
        assert b"Tornillo" in resp.data

    def test_producto_nuevo_get(self, auth_client):
        resp = auth_client.get("/productos/nuevo")
        assert resp.status_code == 200

    def test_producto_nuevo_post(self, auth_client):
        resp = auth_client.post("/productos/nuevo", data={
            "codigo": "P100",
            "descripcion": "Producto Test",
            "um": "UND",
            "familia": "TEST",
            "stock_minimo": "5",
        }, follow_redirects=True)
        assert resp.status_code == 200
        assert b"guardado" in resp.data

    def test_producto_editar(self, auth_client, sample_productos):
        resp = auth_client.post("/productos/editar/1", data={
            "codigo": "P001",
            "descripcion": "Tornillo M8x30 Editado",
            "um": "UND",
            "familia": "FERRETERIA",
            "stock_minimo": "15",
        }, follow_redirects=True)
        assert resp.status_code == 200
        assert b"guardado" in resp.data

    def test_producto_eliminar_sin_movimientos(self, auth_client, sample_productos):
        resp = auth_client.post("/productos/eliminar/3", follow_redirects=True)
        assert resp.status_code == 200
        assert b"eliminado" in resp.data

    def test_producto_eliminar_con_movimientos(self, auth_client, sample_productos, app):
        with app.app_context():
            p = Producto.query.filter_by(codigo="P001").first()
            e = Entrada(producto_id=p.id, cantidad=10)
            db.session.add(e)
            db.session.commit()
        resp = auth_client.post("/productos/eliminar/1", follow_redirects=True)
        assert resp.status_code == 200
        assert b"movimientos asociados" in resp.data

    def test_producto_eliminar_todos_sin_movimientos(self, auth_client, sample_productos, app):
        """Elimina todos los productos que no tienen movimientos."""
        with app.app_context():
            # P001 tiene 1 entrada, P002 no tiene movs, P003 no tiene movs
            p = Producto.query.filter_by(codigo="P001").first()
            e = Entrada(producto_id=p.id, cantidad=10)
            db.session.add(e)
            db.session.commit()
        resp = auth_client.post("/productos/eliminar-todos", follow_redirects=True)
        assert resp.status_code == 200
        assert b"eliminaron" in resp.data
        # P002 y P003 deberían eliminarse, P001 se conserva
        with app.app_context():
            assert Producto.query.filter_by(codigo="P002").first() is None
            assert Producto.query.filter_by(codigo="P003").first() is None
            assert Producto.query.filter_by(codigo="P001").first() is not None

    def test_producto_eliminar_todos_con_todos_con_movimientos(self, auth_client, sample_productos, app):
        """No elimina ningún producto si todos tienen movimientos."""
        with app.app_context():
            for p in Producto.query.all():
                e = Entrada(producto_id=p.id, cantidad=1)
                db.session.add(e)
            db.session.commit()
        resp = auth_client.post("/productos/eliminar-todos", follow_redirects=True)
        assert resp.status_code == 200
        assert b"No hay productos sin movimientos" in resp.data
        with app.app_context():
            assert Producto.query.count() == 3

    def test_producto_eliminar_todos_sin_productos(self, auth_client):
        """No falla si no hay productos en la BD."""
        resp = auth_client.post("/productos/eliminar-todos", follow_redirects=True)
        assert resp.status_code == 200
        assert b"No hay productos sin movimientos" in resp.data


# ===========================================================================
# Tests de Entradas / Salidas manuales
# ===========================================================================

class TestMovimientosManuales:
    def test_entrada_get(self, auth_client):
        resp = auth_client.get("/entradas")
        assert resp.status_code == 200

    def test_entrada_post(self, auth_client, sample_productos):
        resp = auth_client.post("/entradas", data={
            "producto_id": "1",
            "cantidad": "25",
            "zona": "A",
            "ubicacion": "EST-01",
            "alm": "ALM-01",
            "oc": "OC-001",
            "guia_remision": "G-001",
        }, follow_redirects=True)
        assert resp.status_code == 200
        assert b"Entrada registrada" in resp.data

    def test_entrada_cantidad_invalida(self, auth_client, sample_productos):
        resp = auth_client.post("/entradas", data={
            "producto_id": "1",
            "cantidad": "-5",
        }, follow_redirects=True)
        assert resp.status_code == 200
        assert b"mayor a 0" in resp.data

    def test_salida_get(self, auth_client):
        resp = auth_client.get("/salidas")
        assert resp.status_code == 200

    def test_salida_post(self, auth_client, sample_productos):
        resp = auth_client.post("/salidas", data={
            "producto_id": "1",
            "cantidad": "10",
            "nro_vale": "V-001",
            "oi": "OI-001",
            "c_costo": "CC-100",
            "maquina": "M-01",
            "categoria": "GENERAL",
        }, follow_redirects=True)
        assert resp.status_code == 200
        assert b"Salida registrada" in resp.data

    def test_salida_stock_insuficiente(self, auth_client, sample_productos):
        resp = auth_client.post("/salidas", data={
            "producto_id": "3",
            "cantidad": "10",
        }, follow_redirects=True)
        assert resp.status_code == 200
        assert b"insuficiente" in resp.data

    def test_entrada_editar_get(self, auth_client, sample_productos, app):
        with app.app_context():
            p = Producto.query.filter_by(codigo="P001").first()
            e = Entrada(producto_id=p.id, cantidad=20)
            db.session.add(e)
            db.session.commit()
            eid = e.id
        resp = auth_client.get(f"/entradas/editar/{eid}")
        assert resp.status_code == 200
        assert b"Editar Entrada" in resp.data

    def test_entrada_editar_post(self, auth_client, sample_productos, app):
        with app.app_context():
            p = Producto.query.filter_by(codigo="P001").first()
            e = Entrada(producto_id=p.id, cantidad=10)
            db.session.add(e)
            db.session.commit()
            eid = e.id
            stock_anterior = p.stock_actual
        resp = auth_client.post(f"/entradas/editar/{eid}", data={
            "cantidad": "25",
        }, follow_redirects=True)
        assert resp.status_code == 200
        assert b"actualizada" in resp.data or b"Entrada" in resp.data
        with app.app_context():
            p = Producto.query.filter_by(codigo="P001").first()
            # stock debe haber aumentado en 15 (25-10)
            assert p.stock_actual == stock_anterior + 15

    def test_entrada_eliminar(self, auth_client, sample_productos, app):
        with app.app_context():
            p = Producto.query.filter_by(codigo="P001").first()
            stock_antes = p.stock_actual
        # Crear entrada via route (ajusta stock)
        resp = auth_client.post("/entradas", data={
            "producto_id": "1", "cantidad": "15",
        }, follow_redirects=True)
        assert resp.status_code == 200
        with app.app_context():
            p = Producto.query.filter_by(codigo="P001").first()
            assert p.stock_actual == stock_antes + 15
            e = Entrada.query.filter_by(producto_id=p.id).first()
            eid = e.id
        # Eliminar la entrada
        resp = auth_client.post(f"/entradas/eliminar/{eid}", follow_redirects=True)
        assert resp.status_code == 200
        with app.app_context():
            p = Producto.query.filter_by(codigo="P001").first()
            assert p.stock_actual == stock_antes

    def test_entrada_editar_sin_cambios(self, auth_client, sample_productos, app):
        """Editar entrada sin cambiar cantidad no altera stock."""
        with app.app_context():
            p = Producto.query.filter_by(codigo="P001").first()
            stock_antes = p.stock_actual
        resp = auth_client.post("/entradas", data={
            "producto_id": "1", "cantidad": "10",
        }, follow_redirects=True)
        assert resp.status_code == 200
        with app.app_context():
            e = Entrada.query.filter_by(producto_id=1).first()
            assert e is not None
            eid = e.id
        resp = auth_client.post(f"/entradas/editar/{eid}", data={
            "cantidad": "10",
        }, follow_redirects=True)
        assert resp.status_code == 200
        with app.app_context():
            p = Producto.query.filter_by(codigo="P001").first()
            assert p.stock_actual == stock_antes + 10

    def test_salida_editar_get(self, auth_client, sample_productos, app):
        with app.app_context():
            p = Producto.query.filter_by(codigo="P002").first()
            pid = p.id
        resp = auth_client.post("/salidas", data={
            "producto_id": str(pid), "cantidad": "10",
        }, follow_redirects=True)
        assert resp.status_code == 200
        with app.app_context():
            s = Salida.query.filter_by(producto_id=pid).first()
            assert s is not None, "Salida no fue creada"
            sid = s.id
        resp = auth_client.get(f"/salidas/editar/{sid}")
        assert resp.status_code == 200
        assert b"Editar Salida" in resp.data

    def test_salida_editar_post_decremento(self, auth_client, sample_productos, app):
        """Reducir cantidad de salida (devuelve stock)."""
        with app.app_context():
            p = Producto.query.filter_by(codigo="P002").first()
            stock_antes = p.stock_actual
            pid = p.id
        resp = auth_client.post("/salidas", data={
            "producto_id": str(pid), "cantidad": "10",
        }, follow_redirects=True)
        assert resp.status_code == 200
        with app.app_context():
            p = Producto.query.filter_by(codigo="P002").first()
            assert p.stock_actual == stock_antes - 10, f"Stock {p.stock_actual} != {stock_antes - 10}"
            s = Salida.query.filter_by(producto_id=pid).first()
            assert s is not None
            sid = s.id
        resp = auth_client.post(f"/salidas/editar/{sid}", data={
            "cantidad": "5",
        }, follow_redirects=True)
        assert resp.status_code == 200
        with app.app_context():
            p = Producto.query.filter_by(codigo="P002").first()
            assert p.stock_actual == stock_antes - 5

    def test_salida_eliminar(self, auth_client, sample_productos, app):
        with app.app_context():
            p = Producto.query.filter_by(codigo="P002").first()
            stock_antes = p.stock_actual
            pid = p.id
        resp = auth_client.post("/salidas", data={
            "producto_id": str(pid), "cantidad": "5",
        }, follow_redirects=True)
        assert resp.status_code == 200
        with app.app_context():
            p = Producto.query.filter_by(codigo="P002").first()
            assert p.stock_actual == stock_antes - 5
            s = Salida.query.filter_by(producto_id=pid).first()
            assert s is not None
            sid = s.id
        resp = auth_client.post(f"/salidas/eliminar/{sid}", follow_redirects=True)
        assert resp.status_code == 200
        with app.app_context():
            p = Producto.query.filter_by(codigo="P002").first()
            assert p.stock_actual == stock_antes


# ===========================================================================
# Tests de Existencias e Historial
# ===========================================================================

class TestConsultas:
    def test_existencias(self, auth_client, sample_productos):
        resp = auth_client.get("/existencias")
        assert resp.status_code == 200
        assert b"Tornillo" in resp.data

    def test_existencias_search(self, auth_client, sample_productos):
        resp = auth_client.get("/existencias?search=Tuerca")
        assert resp.status_code == 200
        assert b"Tuerca" in resp.data

    def test_existencias_solo_bajo(self, auth_client, sample_productos):
        resp = auth_client.get("/existencias?solo_bajo=1")
        assert resp.status_code == 200

    def test_existencias_familia(self, auth_client, sample_productos):
        resp = auth_client.get("/existencias?familia=FERRETERIA")
        assert resp.status_code == 200

    def test_historial(self, auth_client):
        resp = auth_client.get("/historial")
        assert resp.status_code == 200

    def test_historial_filtros(self, auth_client):
        resp = auth_client.get("/historial?tipo=ENTRADA")
        assert resp.status_code == 200


# ===========================================================================
# Tests de Exportación/Plantilla
# ===========================================================================

class TestExportaciones:
    def test_plantilla(self, auth_client):
        resp = auth_client.get("/productos/plantilla")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.content_type

    def test_exportar(self, auth_client):
        resp = auth_client.get("/productos/exportar")
        assert resp.status_code == 200
        assert "spreadsheetml" in resp.content_type


# ===========================================================================
# Tests de Importar productos desde Excel
# ===========================================================================

class TestImportarProductos:
    def test_importar_get(self, auth_client):
        resp = auth_client.get("/productos/importar")
        assert resp.status_code == 200
        assert b"Importar Productos" in resp.data

    def test_importar_sin_archivo(self, auth_client):
        resp = auth_client.post("/productos/importar", follow_redirects=True)
        assert resp.status_code == 200
        assert b"ning" in resp.data or b"archivo" in resp.data

    def test_importar_archivo_invalido(self, auth_client):
        data = {"archivo": (io.BytesIO(b"not an excel"), "test.txt")}
        resp = auth_client.post("/productos/importar", data=data,
                                content_type="multipart/form-data", follow_redirects=True)
        assert resp.status_code == 200
        assert b".xlsx" in resp.data

    def test_importar_confirmar_sin_tmp(self, auth_client):
        resp = auth_client.post("/productos/importar/confirmar", follow_redirects=True)
        assert resp.status_code == 200
        assert b"no encontrado" in resp.data


# ===========================================================================
# Tests de Importar Entradas desde Excel
# ===========================================================================

class TestImportarEntradas:
    def test_importar_get(self, auth_client):
        resp = auth_client.get("/entradas/importar")
        assert resp.status_code == 200
        assert b"Importar Entradas" in resp.data

    def test_importar_sin_archivo(self, auth_client):
        resp = auth_client.post("/entradas/importar", follow_redirects=True)
        assert resp.status_code == 200
        assert b"ning" in resp.data or b"archivo" in resp.data

    def test_importar_archivo_invalido(self, auth_client):
        data = {"archivo": (io.BytesIO(b"not an excel"), "test.txt")}
        resp = auth_client.post("/entradas/importar", data=data,
                                content_type="multipart/form-data", follow_redirects=True)
        assert resp.status_code == 200
        assert b".xlsx" in resp.data

    def test_importar_confirmar_sin_tmp(self, auth_client):
        resp = auth_client.post("/entradas/importar/confirmar", follow_redirects=True)
        assert resp.status_code == 200
        assert b"no encontrado" in resp.data


# ===========================================================================
# Tests de Importar Salidas desde Excel
# ===========================================================================

class TestImportarSalidas:
    def test_importar_get(self, auth_client):
        resp = auth_client.get("/salidas/importar")
        assert resp.status_code == 200
        assert b"Importar Salidas" in resp.data

    def test_importar_sin_archivo(self, auth_client):
        resp = auth_client.post("/salidas/importar", follow_redirects=True)
        assert resp.status_code == 200
        assert b"ning" in resp.data or b"archivo" in resp.data

    def test_importar_archivo_invalido(self, auth_client):
        data = {"archivo": (io.BytesIO(b"not an excel"), "test.txt")}
        resp = auth_client.post("/salidas/importar", data=data,
                                content_type="multipart/form-data", follow_redirects=True)
        assert resp.status_code == 200
        assert b".xlsx" in resp.data

    def test_importar_confirmar_sin_tmp(self, auth_client):
        resp = auth_client.post("/salidas/importar/confirmar", follow_redirects=True)
        assert resp.status_code == 200
        assert b"no encontrado" in resp.data


# ===========================================================================
# Tests de la función helper _parse_excel_movimiento
# ===========================================================================

class TestParseExcelMovimiento:
    def test_ruta_inexistente(self, app):
        from app.routes import _parse_excel_movimiento
        headers, indices, filas, errores = _parse_excel_movimiento(
            "entrada", "/ruta/inexistente.xlsx"
        )
        assert len(errores) > 0
        assert headers is None
        assert filas is None

    def test_workbook_vacio(self, app):
        import openpyxl
        from app.routes import _parse_excel_movimiento

        wb = openpyxl.Workbook()
        ws = wb.active
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        tmp.close()

        try:
            _, _, _, errores = _parse_excel_movimiento("entrada", tmp.name)
            assert len(errores) > 0
            assert any("vac" in e.lower() for e in errores)
        finally:
            os.unlink(tmp.name)

    def test_parse_entrada_valida(self, app, sample_productos):
        import openpyxl
        from app.routes import _parse_excel_movimiento

        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ["CODIGO", "CANTIDAD", "U.M2", "ZONA", "UBICACION", "ALM", "F.INGRESO", "OC", "G.REMISION"]
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h)
        ws.cell(row=2, column=1, value="P001")
        ws.cell(row=2, column=2, value=25.0)
        ws.cell(row=2, column=3, value="UND")
        ws.cell(row=2, column=4, value="A")
        ws.cell(row=2, column=5, value="EST-01")
        ws.cell(row=2, column=6, value="ALM-01")
        ws.cell(row=2, column=7, value="01/01/2025")
        ws.cell(row=2, column=8, value="OC-001")
        ws.cell(row=2, column=9, value="G-001")

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        tmp.close()

        try:
            _, _, filas, errores = _parse_excel_movimiento("entrada", tmp.name)
            assert len(errores) == 0, f"Errores: {errores}"
            assert len(filas) == 1
            assert filas[0]["codigo"] == "P001"
            assert filas[0]["cantidad"] == 25.0
            assert filas[0]["zona"] == "A"
            assert filas[0]["oc"] == "OC-001"
        finally:
            os.unlink(tmp.name)

    def test_parse_salida_valida(self, app, sample_productos):
        import openpyxl
        from app.routes import _parse_excel_movimiento

        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ["CODIGO", "CANTIDAD", "U.M2", "F. SALIDA", "N° VALE", "OI", "C.COSTO", "MAQUINA", "CATEGORIA"]
        for ci, h in enumerate(headers, 1):
            ws.cell(row=1, column=ci, value=h)
        ws.cell(row=2, column=1, value="P001")
        ws.cell(row=2, column=2, value=5.0)
        ws.cell(row=2, column=3, value="UND")
        ws.cell(row=2, column=4, value="15/01/2025")
        ws.cell(row=2, column=5, value="V-001")
        ws.cell(row=2, column=6, value="OI-001")
        ws.cell(row=2, column=7, value="CC-100")
        ws.cell(row=2, column=8, value="M-01")
        ws.cell(row=2, column=9, value="GENERAL")

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        tmp.close()

        try:
            _, _, filas, errores = _parse_excel_movimiento("salida", tmp.name)
            assert len(errores) == 0, f"Errores: {errores}"
            assert len(filas) == 1
            assert filas[0]["codigo"] == "P001"
            assert filas[0]["cantidad"] == 5.0
            assert filas[0]["nro_vale"] == "V-001"
        finally:
            os.unlink(tmp.name)

    def test_sin_columna_codigo(self, app):
        import openpyxl
        from app.routes import _parse_excel_movimiento

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="FECHA")
        ws.cell(row=1, column=2, value="CANTIDAD")

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        tmp.close()

        try:
            _, _, _, errores = _parse_excel_movimiento("entrada", tmp.name)
            assert len(errores) > 0
            assert "CODIGO" in errores[0]
        finally:
            os.unlink(tmp.name)

    def test_cantidad_cero_ignorada(self, app, sample_productos):
        import openpyxl
        from app.routes import _parse_excel_movimiento

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="CODIGO")
        ws.cell(row=1, column=2, value="CANTIDAD")
        ws.cell(row=2, column=1, value="P001")
        ws.cell(row=2, column=2, value=0.0)
        ws.cell(row=3, column=1, value="P002")
        ws.cell(row=3, column=2, value=-5.0)

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        tmp.close()

        try:
            _, _, filas, errores = _parse_excel_movimiento("entrada", tmp.name)
            assert len(errores) == 0
            assert len(filas) == 0  # all filtered out
        finally:
            os.unlink(tmp.name)


# ===========================================================================
# Tests de Órdenes de Compra
# ===========================================================================

class TestOrdenesCompra:
    """Tests para el CRUD de Órdenes de Compra (FASE 5)."""

    def test_oc_lista_vacia(self, auth_client):
        """Lista de OCs vacía muestra mensaje."""
        resp = auth_client.get("/oc")
        assert resp.status_code == 200
        assert b"No hay" in resp.data

    def test_oc_nuevo_get(self, auth_client):
        """GET del formulario de nueva OC."""
        resp = auth_client.get("/oc/nuevo")
        assert resp.status_code == 200

    def test_oc_nuevo_post(self, auth_client):
        """Crear una OC correctamente."""
        resp = auth_client.post("/oc/nuevo", data={
            "numero": "OC-001",
            "proveedor": "Proveedor SAC",
            "estado": "PENDIENTE",
        }, follow_redirects=True)
        assert resp.status_code == 200
        assert b"OC-001" in resp.data
        assert b"creada" in resp.data

    def test_oc_nuevo_duplicado(self, auth_client):
        """Crear OC con número duplicado debe fallar."""
        auth_client.post("/oc/nuevo", data={
            "numero": "OC-001",
            "proveedor": "Proveedor SAC",
        }, follow_redirects=True)
        resp = auth_client.post("/oc/nuevo", data={
            "numero": "OC-001",
            "proveedor": "Otro Proveedor",
        }, follow_redirects=True)
        assert resp.status_code == 200
        assert b"ya existe" in resp.data

    def test_oc_nuevo_sin_numero(self, auth_client):
        """Crear OC sin número debe fallar."""
        resp = auth_client.post("/oc/nuevo", data={
            "numero": "",
            "proveedor": "Proveedor SAC",
        }, follow_redirects=True)
        assert resp.status_code == 200
        assert b"obligatorio" in resp.data

    def test_oc_editar_get(self, auth_client):
        """GET del formulario de edición de OC."""
        auth_client.post("/oc/nuevo", data={
            "numero": "OC-001", "proveedor": "Proveedor SAC",
        }, follow_redirects=True)
        resp = auth_client.get("/oc/editar/1")
        assert resp.status_code == 200
        assert b"Editar" in resp.data

    def test_oc_editar_post(self, auth_client):
        """Editar OC correctamente."""
        auth_client.post("/oc/nuevo", data={
            "numero": "OC-001", "proveedor": "Proveedor SAC",
        }, follow_redirects=True)
        resp = auth_client.post("/oc/editar/1", data={
            "numero": "OC-001-MOD",
            "proveedor": "Nuevo Proveedor",
            "estado": "PARCIAL",
        }, follow_redirects=True)
        assert resp.status_code == 200
        assert b"actualizada" in resp.data
        assert b"OC-001-MOD" in resp.data

    def test_oc_cerrar(self, auth_client):
        """Cerrar una OC correctamente."""
        auth_client.post("/oc/nuevo", data={
            "numero": "OC-001", "proveedor": "Proveedor SAC",
        }, follow_redirects=True)
        resp = auth_client.post("/oc/cerrar/1", follow_redirects=True)
        assert resp.status_code == 200
        assert b"CERRADA" in resp.data
        assert b"marcada" in resp.data

    def test_oc_eliminar_sin_entradas(self, auth_client, app):
        """Eliminar OC sin entradas vinculadas."""
        auth_client.post("/oc/nuevo", data={
            "numero": "OC-001", "proveedor": "Proveedor SAC",
        }, follow_redirects=True)
        resp = auth_client.post("/oc/eliminar/1", follow_redirects=True)
        assert resp.status_code == 200
        assert b"eliminada" in resp.data

    def test_oc_eliminar_con_entradas(self, auth_client, app, sample_productos):
        """No se puede eliminar OC con entradas vinculadas."""
        auth_client.post("/oc/nuevo", data={
            "numero": "OC-001", "proveedor": "Proveedor SAC",
        }, follow_redirects=True)
        # Crear entrada vinculada a la OC
        auth_client.post("/entradas", data={
            "producto_id": "1", "cantidad": "10", "oc": "OC-001",
        }, follow_redirects=True)
        resp = auth_client.post("/oc/eliminar/1", follow_redirects=True)
        assert resp.status_code == 200
        assert b"No se puede eliminar" in resp.data
        assert b"entrada" in resp.data

    def test_oc_editar_no_existente(self, auth_client):
        """Editar OC que no existe redirige con error."""
        resp = auth_client.get("/oc/editar/999", follow_redirects=True)
        assert resp.status_code == 200
        assert b"no encontrada" in resp.data

    def test_oc_filtro_estado(self, auth_client):
        """Filtrar OCs por estado."""
        auth_client.post("/oc/nuevo", data={"numero": "OC-PEND", "estado": "PENDIENTE"}, follow_redirects=True)
        auth_client.post("/oc/nuevo", data={"numero": "OC-CERR", "estado": "CERRADA"}, follow_redirects=True)
        resp = auth_client.get("/oc?estado=CERRADA")
        assert resp.status_code == 200
        assert b"OC-CERR" in resp.data
        assert b"OC-PEND" not in resp.data

    def test_oc_busqueda_numero(self, auth_client):
        """Buscar OC por número."""
        auth_client.post("/oc/nuevo", data={"numero": "OC-ABC-001"}, follow_redirects=True)
        auth_client.post("/oc/nuevo", data={"numero": "OC-XYZ-002"}, follow_redirects=True)
        resp = auth_client.get("/oc?search=ABC")
        assert resp.status_code == 200
        assert b"OC-ABC-001" in resp.data
        assert b"OC-XYZ-002" not in resp.data

    def test_oc_cambio_estado_con_entrada_parcial(self, auth_client, app, sample_productos):
        """Al registrar una entrada con OC, el estado de la OC NO cambia automáticamente."""
        auth_client.post("/oc/nuevo", data={
            "numero": "OC-001", "proveedor": "Proveedor SAC",
        }, follow_redirects=True)
        auth_client.post("/entradas", data={
            "producto_id": "1", "cantidad": "10", "oc": "OC-001",
        }, follow_redirects=True)
        # El estado debe seguir siendo PENDIENTE (no cambia automáticamente)
        resp = auth_client.get("/oc")
        assert resp.status_code == 200
        assert b"PENDIENTE" in resp.data
