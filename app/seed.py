"""
Script para cargar datos iniciales desde un archivo Excel o generar datos de ejemplo.

Uso:
    python -m app.seed              # lee datos_iniciales.xlsx (si existe) y carga
    python -m app.seed --demo       # genera datos ficticios de demostración
    python -m app.seed --reset      # borra todos los datos y recarga desde cero

El Excel debe tener las hojas: MAESTRA, ENTRADA, SALIDAS
"""

import os
import sys
import argparse
from datetime import datetime, timezone, timedelta
from app import create_app, db
from app.models import User, Producto, Entrada, Salida

app = create_app()

# ---------------------------------------------------------------------------
# Datos de demostración (fallback si no hay Excel)
# ---------------------------------------------------------------------------

PRODUCTOS_DEMO = [
    {"codigo": "P001", "cod_catalogo": "CAT-001", "descripcion": "Tornillo hexagonal M8x30", "um": "UND", "familia": "FERRETERIA", "stock_minimo": 50},
    {"codigo": "P002", "cod_catalogo": "CAT-001", "descripcion": "Tornillo hexagonal M10x40", "um": "UND", "familia": "FERRETERIA", "stock_minimo": 40},
    {"codigo": "P003", "cod_catalogo": "CAT-002", "descripcion": "Tuerca M8", "um": "UND", "familia": "FERRETERIA", "stock_minimo": 60},
    {"codigo": "P004", "cod_catalogo": "CAT-002", "descripcion": "Tuerca M10", "um": "UND", "familia": "FERRETERIA", "stock_minimo": 50},
    {"codigo": "P005", "cod_catalogo": "CAT-003", "descripcion": "Arandela plana M8", "um": "UND", "familia": "FERRETERIA", "stock_minimo": 100},
    {"codigo": "P006", "cod_catalogo": "CAT-004", "descripcion": "Aceite hidráulico ISO 68 - Balde 5GL", "um": "GLN", "familia": "LUBRICANTES", "stock_minimo": 5},
    {"codigo": "P007", "cod_catalogo": "CAT-004", "descripcion": "Grasa multipropósito EP-2 - 1kg", "um": "KG", "familia": "LUBRICANTES", "stock_minimo": 10},
    {"codigo": "P008", "cod_catalogo": "CAT-005", "descripcion": "Rodamiento 6205-2RS", "um": "UND", "familia": "RODAMIENTOS", "stock_minimo": 8},
    {"codigo": "P009", "cod_catalogo": "CAT-005", "descripcion": "Rodamiento 6306-2RS", "um": "UND", "familia": "RODAMIENTOS", "stock_minimo": 6},
    {"codigo": "P010", "cod_catalogo": "CAT-006", "descripcion": "Faja trapezoidal A-40", "um": "UND", "familia": "TRANSMISION", "stock_minimo": 15},
    {"codigo": "P011", "cod_catalogo": "CAT-006", "descripcion": "Faja trapezoidal B-60", "um": "UND", "familia": "TRANSMISION", "stock_minimo": 12},
    {"codigo": "P012", "cod_catalogo": "CAT-007", "descripcion": "Filtro de aceite motor - Modelo X", "um": "UND", "familia": "FILTROS", "stock_minimo": 20},
    {"codigo": "P013", "cod_catalogo": "CAT-007", "descripcion": "Filtro de aire primario", "um": "UND", "familia": "FILTROS", "stock_minimo": 15},
    {"codigo": "P014", "cod_catalogo": "CAT-008", "descripcion": "Cemento Portland Tipo I - Bolsa 42.5kg", "um": "BOL", "familia": "CONSTRUCCION", "stock_minimo": 30},
    {"codigo": "P015", "cod_catalogo": "CAT-008", "descripcion": "Arena fina - m3", "um": "M3", "familia": "CONSTRUCCION", "stock_minimo": 10},
    {"codigo": "P016", "cod_catalogo": "CAT-009", "descripcion": "Pintura esmalte sintético blanco - GL", "um": "GLN", "familia": "PINTURAS", "stock_minimo": 8},
    {"codigo": "P017", "cod_catalogo": "CAT-009", "descripcion": "Thinner acrílico - GL", "um": "GLN", "familia": "PINTURAS", "stock_minimo": 10},
    {"codigo": "P018", "cod_catalogo": "CAT-010", "descripcion": "Cable eléctrico #12 AWG - metro", "um": "M", "familia": "ELECTRICIDAD", "stock_minimo": 200},
    {"codigo": "P019", "cod_catalogo": "CAT-010", "descripcion": "Interruptor simple 15A", "um": "UND", "familia": "ELECTRICIDAD", "stock_minimo": 30},
    {"codigo": "P020", "cod_catalogo": "CAT-011", "descripcion": "Guante de seguridad talla L", "um": "PAR", "familia": "EPP", "stock_minimo": 50},
    {"codigo": "P021", "cod_catalogo": "CAT-011", "descripcion": "Casco de seguridad blanco", "um": "UND", "familia": "EPP", "stock_minimo": 25},
    {"codigo": "P022", "cod_catalogo": "CAT-012", "descripcion": "Soldadura 6011 1/8 pulg - kg", "um": "KG", "familia": "SOLDADURA", "stock_minimo": 20},
    {"codigo": "P023", "cod_catalogo": "CAT-012", "descripcion": "Disco de corte 7 x 1/8 pulg", "um": "UND", "familia": "SOLDADURA", "stock_minimo": 40},
    {"codigo": "P024", "cod_catalogo": "CAT-013", "descripcion": "Tubería PVC 1/2 pulg - 3m", "um": "UND", "familia": "PLOMERIA", "stock_minimo": 20},
    {"codigo": "P025", "cod_catalogo": "CAT-013", "descripcion": "Codo PVC 1/2 pulg 90°", "um": "UND", "familia": "PLOMERIA", "stock_minimo": 30},
]

ZONAS = ["ZONA-A", "ZONA-B", "ZONA-C"]
UBICACIONES = ["EST-01", "EST-02", "EST-03", "RACK-01", "RACK-02", "PISO-01"]


def _generar_entradas_demo(productos):
    """Genera entradas de ejemplo para los productos."""
    entradas = []
    base_fecha = datetime.now(timezone.utc) - timedelta(days=60)
    for i, prod in enumerate(productos):
        # 1-2 entradas por producto
        for j in range(2):
            dias_atras = (i * 3 + j * 5) % 60
            cantidad = (i + 1) * 10.0 + j * 5.0
            entradas.append({
                "producto": prod,
                "cantidad": cantidad,
                "zona": ZONAS[i % len(ZONAS)],
                "ubicacion": UBICACIONES[(i + j) % len(UBICACIONES)],
                "alm": "ALM-01",
                "fecha": base_fecha + timedelta(days=dias_atras),
                "oc": f"OC-{2024000 + i * 10 + j}",
                "guia_remision": f"GR-{1000 + i * 10 + j}",
            })
    return entradas


def _generar_salidas_demo(productos):
    """Genera salidas de ejemplo para los productos."""
    salidas = []
    base_fecha = datetime.now(timezone.utc) - timedelta(days=30)
    for i, prod in enumerate(productos[:15]):  # Solo algunos productos tienen salidas
        for j in range(2):
            dias_atras = (i * 2 + j * 7) % 30
            cantidad = (i + 1) * 2.0 + j
            salidas.append({
                "producto": prod,
                "cantidad": cantidad,
                "fecha": base_fecha + timedelta(days=dias_atras),
                "nro_vale": f"V-{500 + i * 5 + j}",
                "oi": f"OI-{300 + i * 3 + j}",
                "c_costo": f"CC-{100 + i * 2}",
                "maquina": f"MAQ-{10 + i % 5}",
                "categoria": prod.familia,
            })
    return salidas


# ---------------------------------------------------------------------------
# Carga desde Excel
# ---------------------------------------------------------------------------

def _leer_excel(ruta_excel):
    """Lee el archivo Excel y retorna listas de diccionarios."""
    try:
        import openpyxl
    except ImportError:
        print("Error: openpyxl no está instalado. Instálalo con: pip install openpyxl")
        return None, None, None

    wb = openpyxl.load_workbook(ruta_excel, data_only=True)

    productos_data = []
    if "MAESTRA" in wb.sheetnames:
        ws = wb["MAESTRA"]
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = dict(zip(headers, row))
            if vals.get("CODIGO"):
                productos_data.append(vals)
        print(f"  [MAESTRA] {len(productos_data)} productos leídos.")
    else:
        print("  [MAESTRA] Hoja no encontrada en el Excel.")

    entradas_data = []
    if "ENTRADA" in wb.sheetnames:
        ws = wb["ENTRADA"]
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = dict(zip(headers, row))
            if vals.get("CODIGO"):
                entradas_data.append(vals)
        print(f"  [ENTRADA] {len(entradas_data)} entradas leídas.")
    else:
        print("  [ENTRADA] Hoja no encontrada en el Excel.")

    salidas_data = []
    if "SALIDAS" in wb.sheetnames:
        ws = wb["SALIDAS"]
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            vals = dict(zip(headers, row))
            if vals.get("CODIGO"):
                salidas_data.append(vals)
        print(f"  [SALIDAS] {len(salidas_data)} salidas leídas.")
    else:
        print("  [SALIDAS] Hoja no encontrada en el Excel.")

    return productos_data, entradas_data, salidas_data


def _cargar_desde_excel(productos_data, entradas_data, salidas_data):
    """Carga los datos del Excel a la base de datos."""
    # Mapeo de código a Producto
    codigo_a_producto = {}

    for p in productos_data:
        codigo = str(p.get("CODIGO", "")).strip()
        if not codigo:
            continue
        producto = Producto.query.filter_by(codigo=codigo).first()
        if not producto:
            producto = Producto(codigo=codigo)
            db.session.add(producto)
        producto.cod_catalogo = str(p.get("COD. CATALOGO", "")).strip()
        producto.descripcion = str(p.get("DESCRIPCION DEL PRODUCTO", "")).strip()
        producto.um = str(p.get("U.M", "UND")).strip() or "UND"
        producto.familia = str(p.get("FAMILIA", "")).strip()
        codigo_a_producto[codigo] = producto
    db.session.commit()
    print(f"  Productos cargados: {len(codigo_a_producto)}")

    # Cargar entradas
    for e in entradas_data:
        codigo = str(e.get("CODIGO", "")).strip()
        producto = codigo_a_producto.get(codigo)
        if not producto:
            continue
        try:
            cantidad = float(e.get("CANTIDA", 0) or 0)
        except (ValueError, TypeError):
            cantidad = 0
        if cantidad <= 0:
            continue
        fecha_str = str(e.get("F.INGRESO", "")).strip()
        fecha = None
        if fecha_str:
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                try:
                    fecha = datetime.strptime(fecha_str, fmt).replace(tzinfo=timezone.utc)
                    break
                except ValueError:
                    continue
        if fecha is None:
            fecha = datetime.now(timezone.utc)

        entrada = Entrada(
            producto_id=producto.id,
            cantidad=cantidad,
            um=str(e.get("U.M2", producto.um)).strip() or producto.um,
            zona=str(e.get("ZONA", "")).strip(),
            ubicacion=str(e.get("UBICACIÓN", "")).strip(),
            alm=str(e.get("ALM", "ALM-01")).strip(),
            fecha_ingreso=fecha,
            oc=str(e.get("OC", "")).strip(),
            guia_remision=str(e.get("G.REMISION", "")).strip(),
            familia=str(e.get("FAMILIA", producto.familia)).strip() or producto.familia,
        )
        producto.stock_actual += cantidad
        db.session.add(entrada)
    db.session.commit()
    print(f"  Entradas cargadas: {Entrada.query.count()}")

    # Cargar salidas
    for s in salidas_data:
        codigo = str(s.get("CODIGO", "")).strip()
        producto = codigo_a_producto.get(codigo)
        if not producto:
            continue
        try:
            cantidad = float(s.get("CANTIDAD", 0) or 0)
        except (ValueError, TypeError):
            cantidad = 0
        if cantidad <= 0:
            continue
        fecha_str = str(s.get("F. SALIDA", "")).strip()
        fecha = None
        if fecha_str:
            for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
                try:
                    fecha = datetime.strptime(fecha_str, fmt).replace(tzinfo=timezone.utc)
                    break
                except ValueError:
                    continue
        if fecha is None:
            fecha = datetime.now(timezone.utc)

        if cantidad > producto.stock_actual:
            cantidad = producto.stock_actual  # No permitir stock negativo

        salida = Salida(
            producto_id=producto.id,
            cantidad=cantidad,
            um=str(s.get("U.M2", producto.um)).strip() or producto.um,
            fecha_salida=fecha,
            nro_vale=str(s.get("N° VALE", "")).strip(),
            oi=str(s.get("OI", "")).strip(),
            c_costo=str(s.get("C.COSTO", "")).strip(),
            maquina=str(s.get("MAQUINA", "")).strip(),
            categoria=str(s.get("CATEGPRIA", "")).strip(),
        )
        producto.stock_actual -= cantidad
        db.session.add(salida)
    db.session.commit()
    print(f"  Salidas cargadas: {Salida.query.count()}")


# ---------------------------------------------------------------------------
# Carga de datos demo
# ---------------------------------------------------------------------------

def _cargar_demo():
    """Carga productos, entradas y salidas de demostración."""
    for p_data in PRODUCTOS_DEMO:
        existente = Producto.query.filter_by(codigo=p_data["codigo"]).first()
        if not existente:
            producto = Producto(**p_data)
            db.session.add(producto)
    db.session.commit()
    productos = Producto.query.all()
    print(f"  Productos demo cargados: {len(productos)}")

    # Entradas demo
    for e_data in _generar_entradas_demo(productos):
        entrada = Entrada(
            producto_id=e_data["producto"].id,
            cantidad=e_data["cantidad"],
            um=e_data["producto"].um,
            zona=e_data["zona"],
            ubicacion=e_data["ubicacion"],
            alm=e_data["alm"],
            fecha_ingreso=e_data["fecha"],
            oc=e_data["oc"],
            guia_remision=e_data["guia_remision"],
            familia=e_data["producto"].familia,
        )
        e_data["producto"].stock_actual += e_data["cantidad"]
        db.session.add(entrada)
    db.session.commit()
    print(f"  Entradas demo cargadas: {Entrada.query.count()}")

    # Salidas demo
    productos_actualizados = Producto.query.all()
    for s_data in _generar_salidas_demo(productos_actualizados):
        if s_data["cantidad"] > s_data["producto"].stock_actual:
            s_data["cantidad"] = s_data["producto"].stock_actual
        salida = Salida(
            producto_id=s_data["producto"].id,
            cantidad=s_data["cantidad"],
            um=s_data["producto"].um,
            fecha_salida=s_data["fecha"],
            nro_vale=s_data["nro_vale"],
            oi=s_data["oi"],
            c_costo=s_data["c_costo"],
            maquina=s_data["maquina"],
            categoria=s_data["categoria"],
        )
        s_data["producto"].stock_actual -= s_data["cantidad"]
        db.session.add(salida)
    db.session.commit()
    print(f"  Salidas demo cargadas: {Salida.query.count()}")


# ---------------------------------------------------------------------------
# Punto de entrada
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Carga inicial de datos para Almacén")
    parser.add_argument("--demo", action="store_true", help="Cargar datos de demostración")
    parser.add_argument("--reset", action="store_true", help="Borrar todos los datos y recargar")
    parser.add_argument("--excel", type=str, default="datos_iniciales.xlsx", help="Ruta al archivo Excel")
    args = parser.parse_args()

    with app.app_context():
        # Crear usuario administrador
        admin_username = "cponce123.com@gmail.com"
        admin = User.query.filter_by(username=admin_username).first()
        if not admin:
            admin = User(username=admin_username)
            admin_password = os.environ.get("ADMIN_PASSWORD", "Hadrones456%")
            admin.set_password(admin_password)
            db.session.add(admin)
            db.session.commit()
            print(f"Administrador creado: {admin_username}")

        if args.reset:
            print("Eliminando datos existentes...")
            Salida.query.delete()
            Entrada.query.delete()
            Producto.query.delete()
            db.session.commit()
            print("Datos eliminados.")

        # Intentar cargar desde Excel primero
        ruta_excel = args.excel
        if os.path.exists(ruta_excel) and not args.demo:
            print(f"Leyendo archivo Excel: {ruta_excel}")
            prod_data, ent_data, sal_data = _leer_excel(ruta_excel)
            if prod_data:
                _cargar_desde_excel(prod_data, ent_data, sal_data)
                print("Carga desde Excel completada.")
                return

        # Si no hay Excel o se fuerza demo
        if args.demo or not os.path.exists(ruta_excel):
            print("Cargando datos de demostración...")
            _cargar_demo()
            print("Carga de datos demo completada.")
            return

        print("No se encontró el archivo Excel. Usa --demo para datos de demostración.")


if __name__ == "__main__":
    main()
