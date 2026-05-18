import io
import os
import re
from datetime import datetime, timezone, timedelta
from urllib.parse import urlparse
from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file, session, jsonify
from flask_login import login_user, logout_user, login_required, current_user
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import joinedload
from app import db
from app.models import User, Producto, Entrada, Salida, AuditLog, audit_log, Familia, Almacen, OrdenCompra

routes_bp = Blueprint("routes", __name__)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _init_admin_user():
    """Crea el usuario administrador por defecto si no existe.
    
    También elimina el viejo usuario 'admin' si aún existe en la BD.
    La contraseña del admin se configura vía variable de entorno ADMIN_PASSWORD.
    """
    # Eliminar el viejo admin:admin si existe (migración desde versión anterior)
    viejo_admin = User.query.filter_by(username="admin").first()
    if viejo_admin and viejo_admin.username != User.ADMIN_USERNAME:
        db.session.delete(viejo_admin)
        db.session.commit()

    # Crear el nuevo administrador si no existe
    admin_username = User.ADMIN_USERNAME
    if not User.query.filter_by(username=admin_username).first():
        admin_password = os.environ.get("ADMIN_PASSWORD", "Hadrones456%")
        admin = User(username=admin_username)
        admin.set_password(admin_password)
        db.session.add(admin)
        db.session.commit()

    # Crear usuario hsuarez si no existe
    if not User.query.filter_by(username="hsuarez").first():
        hsuarez = User(username="hsuarez")
        hsuarez.set_password("suarez123")
        db.session.add(hsuarez)
        db.session.commit()


def _validar_mime_excel(filepath):
    """Verifica que el archivo sea un Excel .xlsx válido.
    
    Los archivos .xlsx son contenedores ZIP. Verifica el magic header.
    Retorna True si parece válido, False en caso contrario.
    """
    try:
        with open(filepath, "rb") as f:
            header = f.read(4)
        return header == b"PK\x03\x04"  # ZIP header
    except Exception:
        return False


# ---------------------------------------------------------------------------
# Health Check
# ---------------------------------------------------------------------------

@routes_bp.route("/health")
def health():
    """Endpoint de monitoreo para Render."""
    try:
        db.session.execute(db.text("SELECT 1"))
        db_ok = True
    except Exception:
        db_ok = False
    return jsonify({
        "status": "ok" if db_ok else "error",
        "database": "connected" if db_ok else "disconnected",
        "timestamp": datetime.now(timezone.utc).isoformat(),
    })


# ---------------------------------------------------------------------------
# Auth
# ---------------------------------------------------------------------------

@routes_bp.route("/login", methods=["GET", "POST"])
def login():
    _init_admin_user()
    if current_user.is_authenticated:
        return redirect(url_for("routes.dashboard"))

    # --- Rate limiting: 5 intentos, bloqueo 15 minutos ---
    now = datetime.now(timezone.utc)
    login_attempts = session.get("login_attempts", 0)
    blocked_until_str = session.get("login_blocked_until", None)
    blocked_until = None
    if blocked_until_str:
        try:
            blocked_until = datetime.fromisoformat(blocked_until_str)
        except (ValueError, TypeError):
            blocked_until = None

    if blocked_until and now < blocked_until:
        remaining = int((blocked_until - now).total_seconds() // 60)
        flash(
            f"Demasiados intentos. Intenta de nuevo en {remaining} minuto(s).",
            "danger",
        )
        return render_template("login.html")

    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            # Éxito: resetear contadores
            session.pop("login_attempts", None)
            session.pop("login_blocked_until", None)
            login_user(user)
            next_page = request.args.get("next")
            # Validar open redirect: solo URL relativas
            if next_page:
                parsed = urlparse(next_page)
                if parsed.netloc or parsed.scheme:
                    next_page = None  # open redirect malicioso
            flash("Inicio de sesión exitoso.", "success")
            return redirect(next_page or url_for("routes.dashboard"))

        # Fallo: incrementar contador
        login_attempts = session.get("login_attempts", 0) + 1
        session["login_attempts"] = login_attempts
        if login_attempts >= 5:
            blocked_until_dt = now + timedelta(minutes=15)
            session["login_blocked_until"] = blocked_until_dt.isoformat()
            flash(
                "Demasiados intentos fallidos. Bloqueado por 15 minutos.",
                "danger",
            )
        else:
            remaining = 5 - login_attempts
            flash(
                f"Credenciales inválidas. {remaining} intento(s) restante(s).",
                "danger",
            )
    return render_template("login.html")


@routes_bp.route("/logout")
@login_required
def logout():
    logout_user()
    flash("Sesión cerrada correctamente.", "info")
    return redirect(url_for("routes.login"))


@routes_bp.route("/cambiar-contrasena", methods=["GET", "POST"])
@login_required
def cambiar_contrasena():
    """Permite al usuario cambiar su propia contraseña."""
    if request.method == "POST":
        current_password = request.form.get("current_password", "")
        new_password = request.form.get("new_password", "")
        new_password2 = request.form.get("new_password2", "")

        errores = []
        if not current_password:
            errores.append("Debes ingresar tu contraseña actual.")
        elif not current_user.check_password(current_password):
            errores.append("La contraseña actual es incorrecta.")

        if len(new_password) < 4:
            errores.append("La nueva contraseña debe tener al menos 4 caracteres.")
        if new_password != new_password2:
            errores.append("Las contraseñas nuevas no coinciden.")

        if errores:
            for e in errores:
                flash(e, "danger")
            return render_template("cambiar_contrasena.html")

        current_user.set_password(new_password)
        db.session.commit()
        flash("Tu contraseña ha sido cambiada correctamente.", "success")
        return redirect(url_for("routes.dashboard"))

    return render_template("cambiar_contrasena.html")


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------

@routes_bp.route("/")
@login_required
def dashboard():
    total_productos = Producto.query.options(joinedload(Producto.familia_rel)).count()
    total_entradas = Entrada.query.count()
    total_salidas = Salida.query.count()
    total_oc = OrdenCompra.query.count()
    oc_pendientes = OrdenCompra.query.filter(
        OrdenCompra.estado.in_(["PENDIENTE", "PARCIAL"])
    ).count()
    oc_recientes = OrdenCompra.query.order_by(OrdenCompra.created_at.desc()).limit(5).all()
    stock_bajo = Producto.query.options(joinedload(Producto.familia_rel)).filter(
        Producto.stock_minimo > 0,
        Producto.stock_actual <= Producto.stock_minimo
    ).all()

    # Últimos 10 movimientos combinados
    entradas_recientes = Entrada.query.options(joinedload(Entrada.producto)).order_by(Entrada.fecha_ingreso.desc()).limit(10).all()
    salidas_recientes = Salida.query.options(joinedload(Salida.producto)).order_by(Salida.fecha_salida.desc()).limit(10).all()

    movimientos = []
    for e in entradas_recientes:
        movimientos.append({
            "tipo": "ENTRADA",
            "fecha": e.fecha_ingreso,
            "producto": e.producto.descripcion if e.producto else "—",
            "cantidad": e.cantidad,
            "referencia": e.oc or e.guia_remision or "—",
        })
    for s in salidas_recientes:
        movimientos.append({
            "tipo": "SALIDA",
            "fecha": s.fecha_salida,
            "producto": s.producto.descripcion if s.producto else "—",
            "cantidad": s.cantidad,
            "referencia": s.nro_vale or s.oi or "—",
        })
    movimientos.sort(key=lambda m: m["fecha"], reverse=True)
    movimientos = movimientos[:10]

    return render_template(
        "index.html",
        total_productos=total_productos,
        total_entradas=total_entradas,
        total_salidas=total_salidas,
        total_oc=total_oc,
        oc_pendientes=oc_pendientes,
        oc_recientes=oc_recientes,
        stock_bajo=stock_bajo,
        movimientos=movimientos,
    )


# ---------------------------------------------------------------------------
# Productos CRUD
# ---------------------------------------------------------------------------

@routes_bp.route("/productos")
@login_required
def productos():
    search = request.args.get("search", "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = 15

    query = Producto.query.options(joinedload(Producto.familia_rel))
    if search:
        query = query.filter(
            Producto.descripcion.ilike(f"%{search}%")
            | Producto.codigo.ilike(f"%{search}%")
            | Producto.familia.ilike(f"%{search}%")
        )
    query = query.order_by(Producto.codigo.asc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    return render_template("productos.html", pagination=pagination, search=search)


@routes_bp.route("/productos/nuevo", methods=["GET", "POST"])
@login_required
def producto_nuevo():
    return _producto_form()


@routes_bp.route("/productos/editar/<int:producto_id>", methods=["GET", "POST"])
@login_required
def producto_editar(producto_id):
    producto = db.session.get(Producto, producto_id)
    if not producto:
        flash("Producto no encontrado.", "danger")
        return redirect(url_for("routes.productos"))
    return _producto_form(producto)


@routes_bp.route("/productos/eliminar/<int:producto_id>", methods=["POST"])
@login_required
def producto_eliminar(producto_id):
    producto = db.session.get(Producto, producto_id)
    if not producto:
        flash("Producto no encontrado.", "danger")
        return redirect(url_for("routes.productos"))
    if producto.entradas.count() > 0 or producto.salidas.count() > 0:
        flash("No se puede eliminar: el producto tiene movimientos asociados.", "danger")
        return redirect(url_for("routes.productos"))
    db.session.delete(producto)
    db.session.commit()
    flash("Producto eliminado correctamente.", "success")
    return redirect(url_for("routes.productos"))


@routes_bp.route("/productos/eliminar-todos", methods=["POST"])
@login_required
def producto_eliminar_todos():
    """Elimina todos los productos que NO tengan movimientos asociados."""
    productos_con_movimientos = set()
    for e in db.session.query(Entrada.producto_id).distinct().all():
        productos_con_movimientos.add(e[0])
    for s in db.session.query(Salida.producto_id).distinct().all():
        productos_con_movimientos.add(s[0])

    productos_a_eliminar = Producto.query.options(joinedload(Producto.familia_rel)).filter(
        ~Producto.id.in_(productos_con_movimientos)
    ).all() if productos_con_movimientos else Producto.query.options(joinedload(Producto.familia_rel)).all()

    count = len(productos_a_eliminar)
    if count == 0:
        flash("No hay productos sin movimientos para eliminar.", "info")
        return redirect(url_for("routes.productos"))

    for p in productos_a_eliminar:
        db.session.delete(p)
    db.session.commit()
    flash(f"Se eliminaron {count} producto(s) sin movimientos.", "success")
    return redirect(url_for("routes.productos"))


# ---------------------------------------------------------------------------
# Importar / Exportar productos en Excel
# ---------------------------------------------------------------------------

HEADERS_MAESTRA = [
    "CODIGO", "COD. CATALOGO", "DESCRIPCION DEL PRODUCTO", "U.M2", "FAMILIA",
    "REVISADO", "STOCK MINIMO",
]

COL_MAP = {
    "CODIGO": "codigo",
    "COD. CATALOGO": "cod_catalogo",
    "CATALOGO": "cod_catalogo",
    "DESCRIPCION DEL PRODUCTO": "descripcion",
    "DESCRIPCION": "descripcion",
    "DESCRIPCIÓN": "descripcion",
    "PRODUCTO": "descripcion",
    "U.M": "um",
    "U.M2": "um",
    "UM": "um",
    "U.M.": "um",
    "FAMILIA": "familia",
    "REVISADO": "revisado",
    "REVISIÓN": "revisado",
    "REVISION": "revisado",
    "STOCK MINIMO": "stock_minimo",
    "STOCK MÍNIMO": "stock_minimo",
    "STOCK_MINIMO": "stock_minimo",
}

COL_MAP_ENTRADA = {
    "CODIGO": "codigo",
    "COD. CATALOGO": "cod_catalogo",
    "CATALOGO": "cod_catalogo",
    "DESCRIPCION DEL PRODUCTO": "descripcion",
    "DESCRIPCION": "descripcion",
    "DESCRIPCIÓN": "descripcion",
    "CANTIDAD": "cantidad",
    "CANTIDA": "cantidad",
    "U.M2": "um",
    "U.M": "um",
    "UM": "um",
    "ZONA": "zona",
    "UBICACIÓN": "ubicacion",
    "UBICACION": "ubicacion",
    "ALM": "alm",
    "F.INGRESO": "fecha",
    "FECHA": "fecha",
    "FECHA INGRESO": "fecha",
    "OC": "oc",
    "G.REMISION": "guia",
    "GUIA": "guia",
    "GUIA REMISION": "guia",
    "FAMILIA": "familia",
}

COL_MAP_SALIDA = {
    "CODIGO": "codigo",
    "COD. CATALOGO": "cod_catalogo",
    "CATALOGO": "cod_catalogo",
    "DESCRIPCION DEL PRODUCTO": "descripcion",
    "DESCRIPCION": "descripcion",
    "DESCRIPCIÓN": "descripcion",
    "CANTIDAD": "cantidad",
    "U.M2": "um",
    "U.M": "um",
    "UM": "um",
    "F. SALIDA": "fecha",
    "FECHA": "fecha",
    "FECHA SALIDA": "fecha",
    "N° VALE": "nro_vale",
    "NRO VALE": "nro_vale",
    "VALE": "nro_vale",
    "OI": "oi",
    "C.COSTO": "c_costo",
    "C COSTO": "c_costo",
    "COSTO": "c_costo",
    "MAQUINA": "maquina",
    "CATEGORIA": "categoria",
    "CATEGPRIA": "categoria",
}

# Límites máximos según modelos.py
FIELD_MAXLEN = {
    "codigo": 50,
    "cod_catalogo": 50,
    "descripcion": 300,
    "um": 20,
    "familia": 100,
    "zona": 50,
    "ubicacion": 100,
    "alm": 50,
    "oc": 50,
    "guia_remision": 50,
    "nro_vale": 50,
    "oi": 50,
    "c_costo": 100,
    "maquina": 100,
    "categoria": 100,
}

# ---------------------------------------------------------------------------
# Helpers para Excel
# ---------------------------------------------------------------------------


def _excel_val(row, col_indices, col_name):
    """Extraer valor de una celda Excel, normalizado a string."""
    idx = col_indices.get(col_name)
    if idx is None or idx >= len(row):
        return ""
    v = row[idx]
    if v is None:
        return ""
    # Si es datetime, formatear como fecha
    if isinstance(v, (datetime,)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)):
        # Convertir número a string sin decimales si es entero
        if v == int(v):
            return str(int(v))
        return str(v).replace(",", ".")
    return str(v).strip()


def _sanitize_field(valor, field_name):
    """Truncar un campo a su longitud máxima definida."""
    maxlen = FIELD_MAXLEN.get(field_name, 300)
    return str(valor).strip()[:maxlen] if valor else ""


def _make_excel_workbook():
    """Crear workbook con openpyxl y devolver hoja activa."""
    return openpyxl.Workbook()


# ---------------------------------------------------------------------------
# Ruta: Descargar plantilla
# ---------------------------------------------------------------------------

@routes_bp.route("/productos/plantilla")
@login_required
def producto_plantilla():
    """Descargar plantilla Excel vacía con estructura MAESTRA."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MAESTRA"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")

    for col_idx, header in enumerate(HEADERS_MAESTRA, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    widths = [12, 14, 50, 10, 20, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Fila de ejemplo
    ejemplo = ["(Ej: P001)", "(Ej: CAT-001)", "(Ej: Tornillo M8x30)",
               "(Ej: UND)", "(Ej: FERRETERIA)", "(Ej: 10)"]
    for col_idx, val in enumerate(ejemplo, 1):
        cell = ws.cell(row=2, column=col_idx, value=val)
        cell.font = Font(italic=True, color="888888")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="plantilla_maestra.xlsx",
    )


# ---------------------------------------------------------------------------
# API: Búsqueda en vivo de productos (JSON)
# ---------------------------------------------------------------------------

@routes_bp.route("/api/productos")
@login_required
def api_productos():
    """Retorna JSON con productos que coinciden con la búsqueda."""
    q = request.args.get("q", "").strip()
    limit = request.args.get("limit", 20, type=int)

    query = Producto.query.options(joinedload(Producto.familia_rel))
    if q:
        like = f"%{q}%"
        query = query.filter(
            Producto.codigo.ilike(like)
            | Producto.descripcion.ilike(like)
            | Producto.cod_catalogo.ilike(like)
            | Producto.familia.ilike(like)
        )
    productos = query.order_by(Producto.descripcion.asc()).limit(limit).all()

    return {
        "results": [
            {
                "id": p.id,
                "codigo": p.codigo,
                "cod_catalogo": p.cod_catalogo or "",
                "descripcion": p.descripcion,
                "um": p.um,
                "familia": p.familia or "",
                "stock_actual": p.stock_actual,
                "stock_minimo": p.stock_minimo,
            }
            for p in productos
        ],
        "total": len(productos),
    }


HEADERS_PLANTILLA_ENTRADA = [
    "CODIGO", "COD. CATALOGO", "DESCRIPCION DEL PRODUCTO", "CANTIDA",
    "U.M2", "ZONA", "UBICACIÓN", "ALM", "F.INGRESO", "OC", "G.REMISION", "FAMILIA",
]

HEADERS_PLANTILLA_SALIDA = [
    "CODIGO", "COD. CATALOGO", "DESCRIPCION DEL PRODUCTO", "CANTIDAD",
    "U.M2", "F. SALIDA", "N° VALE", "OI", "C.COSTO", "MAQUINA", "CATEGORIA",
]


def _generar_plantilla(headers, titulo, nombre_archivo, ejemplos):
    """Helper para generar una plantilla Excel descargable."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = titulo

    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    ha = Alignment(horizontal="center", vertical="center")

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ha

    widths = [max(12, len(h) + 2) for h in headers]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = min(w, 40)

    if ejemplos:
        for ci, val in enumerate(ejemplos, 1):
            c = ws.cell(row=2, column=ci, value=val)
            c.font = Font(italic=True, color="888888")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output, None


@routes_bp.route("/entradas/plantilla")
@login_required
def entradas_plantilla():
    """Descargar plantilla Excel para importar entradas."""
    ejemplos = ["(Ej: P001)", "(Ej: CAT-001)", "(Ej: Tornillo M8x30)", "(Ej: 10)",
                "(Ej: UND)", "(Ej: ZONA-A)", "(Ej: EST-01)", "(Ej: ALM-01)",
                "(Ej: 01/01/2025)", "(Ej: OC-001)", "(Ej: GR-001)", "(Ej: FERRETERIA)"]
    output, error = _generar_plantilla(
        HEADERS_PLANTILLA_ENTRADA, "ENTRADA", "plantilla_entrada.xlsx", ejemplos
    )
    if error:
        flash(error, "danger")
        return redirect(url_for("routes.entradas_importar"))
    return send_file(output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name="plantilla_entrada.xlsx")


@routes_bp.route("/salidas/plantilla")
@login_required
def salidas_plantilla():
    """Descargar plantilla Excel para importar salidas."""
    ejemplos = ["(Ej: P001)", "(Ej: CAT-001)", "(Ej: Tuerca M8)", "(Ej: 5)",
                "(Ej: UND)", "(Ej: 15/01/2025)", "(Ej: V-001)", "(Ej: OI-001)",
                "(Ej: CC-100)", "(Ej: MAQ-01)", "(Ej: GENERAL)"]
    output, error = _generar_plantilla(
        HEADERS_PLANTILLA_SALIDA, "SALIDA", "plantilla_salida.xlsx", ejemplos
    )
    if error:
        flash(error, "danger")
        return redirect(url_for("routes.salidas_importar"))
    return send_file(output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name="plantilla_salida.xlsx")


# ---------------------------------------------------------------------------
# Ruta: Importar productos desde Excel (con previsualización)
# ---------------------------------------------------------------------------

MAX_IMPORT_ROWS = 5000
"""Número máximo de filas permitidas en una importación."""


def _parse_excel(file):
    """Parsea un archivo Excel y devuelve (headers, col_indices, filas, errores).

    ``file`` puede ser una ruta (str) o un objeto file-like.
    """
    errores_prev = []

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception:
        errores_prev.append("El archivo no es un Excel válido o está corrupto.")
        return None, None, None, errores_prev

    # Leer encabezados
    first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if not first_row or not any(c is not None for c in first_row[0]):
        errores_prev.append("El archivo Excel está vacío.")
        return None, None, None, errores_prev

    headers_raw = []
    for cell_val in first_row[0]:
        s = str(cell_val).strip().upper() if cell_val is not None else ""
        headers_raw.append(s)

    col_indices = {}
    for i, h in enumerate(headers_raw):
        for key, mapped in COL_MAP.items():
            if key.upper() == h:
                col_indices[mapped] = i
                break

    if "codigo" not in col_indices:
        cols = ", ".join(h for h in headers_raw if h)
        errores_prev.append(
            f"No se encontró columna 'CODIGO'. Columnas del archivo: {cols}" if cols
            else "No se encontró la columna 'CODIGO'."
        )
        return headers_raw, None, None, errores_prev

    # Contar filas totales ANTES de procesar
    total_rows = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True)
                     if any(c is not None for c in _))
    if total_rows > MAX_IMPORT_ROWS:
        errores_prev.append(
            f"El archivo tiene {total_rows} filas con datos. "
            f"Máximo permitido: {MAX_IMPORT_ROWS}."
        )
        return headers_raw, col_indices, None, errores_prev

    # Leer filas
    filas = []
    codigos_vistos = set()
    # Obtener todos los códigos existentes de una sola vez (rendimiento)
    codigos_existentes = {r[0] for r in db.session.query(Producto.codigo).all()}

    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None for cell in excel_row):
            continue

        codigo = _excel_val(excel_row, col_indices, "codigo").upper()
        if not codigo:
            continue

        descripcion = _excel_val(excel_row, col_indices, "descripcion")


        # Detectar duplicados internos
        if codigo in codigos_vistos:
            continue
        codigos_vistos.add(codigo)

        cod_catalogo = _excel_val(excel_row, col_indices, "cod_catalogo")
        um = _excel_val(excel_row, col_indices, "um").upper() or "UND"
        familia = _excel_val(excel_row, col_indices, "familia")

        revisado = "Por Revisar"
        if "revisado" in col_indices:
            revisado_val = _excel_val(excel_row, col_indices, "revisado")
            revisado = revisado_val if revisado_val else "Por Revisar"

        stock_minimo = 0.0
        if "stock_minimo" in col_indices:
            raw = _excel_val(excel_row, col_indices, "stock_minimo")
            try:
                stock_minimo = max(0.0, float(raw.replace(",", ".")))
            except (ValueError, AttributeError):
                stock_minimo = 0.0

        filas.append({
            "codigo": codigo,
            "cod_catalogo": cod_catalogo,
            "descripcion": descripcion,
            "um": um,
            "familia": familia,
            "revisado": revisado,
            "stock_minimo": stock_minimo,
            "accion": "ACTUALIZAR" if codigo in codigos_existentes else "CREAR",
        })

    return headers_raw, col_indices, filas, errores_prev


@routes_bp.route("/productos/importar", methods=["GET", "POST"])
@login_required
def producto_importar():
    """Importar productos desde archivo Excel.

    Flujo de dos pasos:
      1. POST sin confirmar → parsea el archivo, muestra vista previa.
      2. POST con confirmar  → ejecuta la importación real.
    """
    # Manejo de peticiones
    import tempfile as _tempfile

    if request.method == "POST":
        # --- Validación del archivo ---
        if "archivo" not in request.files:
            flash("No se seleccionó ningún archivo.", "danger")
            return redirect(url_for("routes.producto_importar"))

        file = request.files["archivo"]
        if not file or file.filename == "":
            flash("No se seleccionó ningún archivo.", "danger")
            return redirect(url_for("routes.producto_importar"))

        # Solo aceptar .xlsx (openpyxl NO soporta .xls antiguo)
        if not file.filename.lower().endswith(".xlsx"):
            flash("Solo se aceptan archivos .xlsx (Excel moderno). "
                  "Si tu archivo es .xls, ábrelo en Excel y guárdalo como .xlsx.", "warning")
            return redirect(url_for("routes.producto_importar"))

        # Guardar el archivo en un temporal para poder releerlo en confirmación
        tmp_file = _tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        try:
            file.save(tmp_file.name)
            tmp_path = tmp_file.name
        except Exception:
            tmp_file.close()
            try:
                os.unlink(tmp_file.name)
            except OSError:
                pass
            raise
        finally:
            tmp_file.close()

        # Validar que sea un Excel real (no un archivo con extensión falsa)
        if not _validar_mime_excel(tmp_path):
            flash("El archivo no es un archivo Excel .xlsx válido.", "danger")
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
            return redirect(url_for("routes.producto_importar"))

        try:
            headers_raw, col_indices, filas, errores_prev = _parse_excel(tmp_path)
        except Exception:
            errores_prev = ["Error inesperado al leer el archivo."]
            headers_raw = None
            filas = None

        if errores_prev:
            # Limpiar temp
            try:
                os.remove(tmp_path)
            except OSError:
                pass
            for e in errores_prev:
                flash(e, "danger")
            return redirect(url_for("routes.producto_importar"))

        if not filas:
            try:
                os.remove(tmp_path)
            except OSError:
                pass
            flash("No se encontraron filas válidas para importar.", "warning")
            return redirect(url_for("routes.producto_importar"))

        # Mostrar vista previa
        headers_mostrar = [h for h in headers_raw if h]
        return render_template(
            "producto_importar.html",
            preview=True,
            tmp_path=tmp_path,
            headers=headers_mostrar,
            filas=filas[:20],  # mostrar primeras 20 filas
            total_filas=len(filas),
            nombre_archivo=file.filename,
        )

    # GET → formulario normal
    return render_template("producto_importar.html")


# ---------------------------------------------------------------------------
# Ruta: Confirmar importación (POST separado para claridad)
# ---------------------------------------------------------------------------

@routes_bp.route("/productos/importar/confirmar", methods=["POST"])
@login_required
def producto_importar_confirmar():
    """Ejecutar la importación después de la previsualización."""
    tmp_path = request.form.get("tmp_path", "").strip()

    if not tmp_path or not os.path.exists(tmp_path):
        flash("Archivo temporal no encontrado. Vuelve a seleccionar el archivo.", "danger")
        return redirect(url_for("routes.producto_importar"))

    try:
        headers_raw, col_indices, filas, errores_prev = _parse_excel(tmp_path)
    except Exception:
        errores_prev = ["Error inesperado al leer el archivo temporal."]
        filas = None

    if errores_prev:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        for e in errores_prev:
            flash(e, "danger")
        return redirect(url_for("routes.producto_importar"))

    if not filas:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        flash("No se encontraron filas válidas para importar.", "warning")
        return redirect(url_for("routes.producto_importar"))

    # --- Ejecutar importación en transacción atómica ---
    creados = []
    actualizados = []
    errores_db = []

    try:
        # Cargar productos existentes de una sola vez
        codigos_a_importar = [f["codigo"] for f in filas]
        existentes_map = {
            p.codigo: p
            for p in Producto.query.options(joinedload(Producto.familia_rel)).filter(Producto.codigo.in_(codigos_a_importar)).all()
        }

        for f in filas:
            codigo = _sanitize_field(f["codigo"], "codigo")
            descripcion = _sanitize_field(f["descripcion"], "descripcion")
            cod_catalogo = _sanitize_field(f["cod_catalogo"], "cod_catalogo")
            um = _sanitize_field(f["um"], "um") or "UND"
            familia = _sanitize_field(f["familia"], "familia")

            producto = existentes_map.get(codigo)
            if producto:
                producto.cod_catalogo = cod_catalogo or producto.cod_catalogo
                producto.descripcion = descripcion or producto.descripcion
                producto.um = um or producto.um
                producto.familia = familia or producto.familia
                producto.stock_minimo = f["stock_minimo"]
                if f.get("revisado"):
                    producto.revisado = f["revisado"]
                actualizados.append(codigo)
            else:
                producto = Producto(
                    codigo=codigo, cod_catalogo=cod_catalogo,
                    descripcion=descripcion or "", um=um, familia=familia,
                    stock_minimo=f["stock_minimo"],
                    revisado=f.get("revisado", "Por Revisar"),
                )
                db.session.add(producto)
                creados.append(codigo)

        db.session.commit()

        # Limpiar temp
        try:
            os.remove(tmp_path)
        except OSError:
            pass

        # Reporte detallado
        partes = []
        if creados:
            partes.append(f"✅ {len(creados)} creados")
        if actualizados:
            partes.append(f"📝 {len(actualizados)} actualizados")
        if errores_db:
            partes.append(f"⚠️ {len(errores_db)} errores")

        msg = "Importación completada: " + ", ".join(partes)

        if errores_db:
            msg += ". Detalles: " + "; ".join(errores_db[:5])
            if len(errores_db) > 5:
                msg += f" y {len(errores_db) - 5} más."
            flash(msg, "warning")
        else:
            flash(msg, "success")

        return redirect(url_for("routes.productos"))

    except Exception as exc:
        db.session.rollback()
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        flash(
            f"Error de base de datos durante la importación: {exc}. "
            "No se guardaron cambios. Verifica que los datos sean válidos.",
            "danger",
        )
        return redirect(url_for("routes.producto_importar"))


# ---------------------------------------------------------------------------
# Helper: Parsear Excel de movimientos (entradas / salidas)
# ---------------------------------------------------------------------------


def _parse_excel_movimiento(tipo, file):
    """Parsea un archivo Excel de movimientos.

    ``tipo`` es 'entrada' o 'salida'.
    ``file`` puede ser ruta (str) o file-like.
    Retorna (headers_raw, col_indices, filas, errores).
    """
    errores = []
    col_map = COL_MAP_ENTRADA if tipo == "entrada" else COL_MAP_SALIDA

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        ws = wb.active
    except Exception:
        errores.append("El archivo no es un Excel válido o está corrupto.")
        return None, None, None, errores

    first_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    if not first_row or not any(c is not None for c in first_row[0]):
        errores.append("El archivo Excel está vacío.")
        return None, None, None, errores

    headers_raw = []
    for cell_val in first_row[0]:
        s = str(cell_val).strip().upper() if cell_val is not None else ""
        headers_raw.append(s)

    col_indices = {}
    for i, h in enumerate(headers_raw):
        for key, mapped in col_map.items():
            if key.upper() == h:
                col_indices[mapped] = i
                break

    if "codigo" not in col_indices:
        cols = ", ".join(h for h in headers_raw if h)
        errores.append(
            f"No se encontró columna 'CODIGO'. Columnas del archivo: {cols}" if cols
            else "No se encontró la columna 'CODIGO'."
        )
        return headers_raw, None, None, errores

    total_rows = sum(
        1 for _ in ws.iter_rows(min_row=2, values_only=True)
        if any(c is not None for c in _)
    )
    if total_rows > MAX_IMPORT_ROWS:
        errores.append(
            f"El archivo tiene {total_rows} filas con datos. "
            f"Máximo permitido: {MAX_IMPORT_ROWS}."
        )
        return headers_raw, col_indices, None, errores

    # Parsear fechas con varios formatos
    def _parse_fecha(val):
        if isinstance(val, datetime):
            return val
        if not val:
            return None
        s = str(val).strip()
        for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%y"):
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
        return None

    filas = []
    for excel_row in ws.iter_rows(min_row=2, values_only=True):
        if not any(cell is not None for cell in excel_row):
            continue

        codigo = _excel_val(excel_row, col_indices, "codigo").upper()
        if not codigo:
            continue

        cantidad = 0.0
        if "cantidad" in col_indices:
            raw = _excel_val(excel_row, col_indices, "cantidad")
            try:
                cantidad = max(0.0, float(raw.replace(",", ".")))
            except (ValueError, AttributeError):
                cantidad = 0.0

        if cantidad <= 0:
            continue

        um = _excel_val(excel_row, col_indices, "um").upper() if "um" in col_indices else ""
        fecha = _parse_fecha(
            _excel_val(excel_row, col_indices, "fecha") if "fecha" in col_indices else ""
        )

        fila = {
            "codigo": codigo,
            "cantidad": cantidad,
            "um": um,
            "fecha": fecha,
        }

        # Campos comunes: catálogo y descripción (opcionales)
        fila["cod_catalogo"] = _sanitize_field(
            _excel_val(excel_row, col_indices, "cod_catalogo") if "cod_catalogo" in col_indices else "", "cod_catalogo"
        )
        fila["descripcion"] = _sanitize_field(
            _excel_val(excel_row, col_indices, "descripcion") if "descripcion" in col_indices else "", "descripcion"
        )

        if tipo == "entrada":
            fila["zona"] = _sanitize_field(
                _excel_val(excel_row, col_indices, "zona") if "zona" in col_indices else "", "zona"
            )
            fila["ubicacion"] = _sanitize_field(
                _excel_val(excel_row, col_indices, "ubicacion") if "ubicacion" in col_indices else "", "ubicacion"
            )
            fila["alm"] = _sanitize_field(
                _excel_val(excel_row, col_indices, "alm") if "alm" in col_indices else "", "alm"
            ) or "ALM-01"
            fila["oc"] = _sanitize_field(
                _excel_val(excel_row, col_indices, "oc") if "oc" in col_indices else "", "oc"
            )
            fila["guia"] = _sanitize_field(
                _excel_val(excel_row, col_indices, "guia") if "guia" in col_indices else "", "guia_remision"
            )
            fila["familia"] = _sanitize_field(
                _excel_val(excel_row, col_indices, "familia") if "familia" in col_indices else "", "familia"
            )
        else:
            fila["nro_vale"] = _sanitize_field(
                _excel_val(excel_row, col_indices, "nro_vale") if "nro_vale" in col_indices else "", "nro_vale"
            )
            fila["oi"] = _sanitize_field(
                _excel_val(excel_row, col_indices, "oi") if "oi" in col_indices else "", "oi"
            )
            fila["c_costo"] = _sanitize_field(
                _excel_val(excel_row, col_indices, "c_costo") if "c_costo" in col_indices else "", "c_costo"
            )
            fila["maquina"] = _sanitize_field(
                _excel_val(excel_row, col_indices, "maquina") if "maquina" in col_indices else "", "maquina"
            )
            fila["categoria"] = _sanitize_field(
                _excel_val(excel_row, col_indices, "categoria") if "categoria" in col_indices else "", "categoria"
            )

        filas.append(fila)

    return headers_raw, col_indices, filas, errores


# ---------------------------------------------------------------------------
# Importar Entradas desde Excel
# ---------------------------------------------------------------------------

@routes_bp.route("/entradas/importar", methods=["GET", "POST"])
@login_required
def entradas_importar():
    import tempfile as _tempfile

    if request.method == "POST":
        if "archivo" not in request.files:
            flash("No se seleccionó ningún archivo.", "danger")
            return redirect(url_for("routes.entradas_importar"))

        file = request.files["archivo"]
        if not file or file.filename == "":
            flash("No se seleccionó ningún archivo.", "danger")
            return redirect(url_for("routes.entradas_importar"))

        if not file.filename.lower().endswith(".xlsx"):
            flash("Solo se aceptan archivos .xlsx (Excel moderno).", "warning")
            return redirect(url_for("routes.entradas_importar"))

        tmp_file = _tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        try:
            file.save(tmp_file.name)
            tmp_path = tmp_file.name
        except Exception:
            tmp_file.close()
            try:
                os.unlink(tmp_file.name)
            except OSError:
                pass
            raise
        finally:
            tmp_file.close()

        # Validar que sea un Excel real (no un archivo con extensión falsa)
        if not _validar_mime_excel(tmp_path):
            flash("El archivo no es un archivo Excel .xlsx válido.", "danger")
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
            return redirect(url_for("routes.entradas_importar"))

        try:
            headers_raw, col_indices, filas, errores_prev = _parse_excel_movimiento("entrada", tmp_path)
        except Exception:
            errores_prev = ["Error inesperado al leer el archivo."]
            headers_raw = None
            filas = None

        if errores_prev:
            try:
                os.remove(tmp_path)
            except OSError:
                pass
            for e in errores_prev:
                flash(e, "danger")
            return redirect(url_for("routes.entradas_importar"))

        if not filas:
            try:
                os.remove(tmp_path)
            except OSError:
                pass
            flash("No se encontraron filas válidas para importar.", "warning")
            return redirect(url_for("routes.entradas_importar"))

        headers_mostrar = [h for h in headers_raw if h]
        return render_template(
            "entradas_importar.html",
            preview=True,
            tmp_path=tmp_path,
            headers=headers_mostrar,
            filas=filas[:20],
            total_filas=len(filas),
            nombre_archivo=file.filename,
        )

    return render_template("entradas_importar.html")


@routes_bp.route("/entradas/importar/confirmar", methods=["POST"])
@login_required
def entradas_importar_confirmar():
    tmp_path = request.form.get("tmp_path", "").strip()
    if not tmp_path or not os.path.exists(tmp_path):
        flash("Archivo temporal no encontrado. Vuelve a seleccionar el archivo.", "danger")
        return redirect(url_for("routes.entradas_importar"))

    try:
        headers_raw, col_indices, filas, errores_prev = _parse_excel_movimiento("entrada", tmp_path)
    except Exception:
        errores_prev = ["Error inesperado al leer el archivo temporal."]
        filas = None

    if errores_prev:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        for e in errores_prev:
            flash(e, "danger")
        return redirect(url_for("routes.entradas_importar"))

    if not filas:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        flash("No se encontraron filas válidas para importar.", "warning")
        return redirect(url_for("routes.entradas_importar"))

    insertadas = []
    errores = []

    try:
        for f in filas:
            producto = Producto.query.options(joinedload(Producto.familia_rel)).filter_by(codigo=f["codigo"]).first()
            if not producto:
                errores.append(f"'{f['codigo']}': producto no encontrado en BD")
                continue

            entrada = Entrada(
                producto_id=producto.id,
                cantidad=f["cantidad"],
                um=f.get("um") or producto.um,
                zona=f.get("zona", ""),
                ubicacion=f.get("ubicacion", ""),
                alm=f.get("alm", "ALM-01"),
                oc=f.get("oc", ""),
                guia_remision=f.get("guia", ""),
                familia=f.get("familia", ""),
                fecha_ingreso=f.get("fecha") or datetime.now(timezone.utc),
            )
            producto.stock_actual += f["cantidad"]
            db.session.add(entrada)
            insertadas.append(f["codigo"])

        db.session.commit()
        try:
            os.remove(tmp_path)
        except OSError:
            pass

        partes = [f"✅ {len(insertadas)} entradas registradas"]
        if errores:
            partes.append(f"⚠️ {len(errores)} errores")
        msg = "Importación completada: " + ", ".join(partes)
        if errores:
            msg += ". Detalles: " + "; ".join(errores[:5])
            if len(errores) > 5:
                msg += f" y {len(errores) - 5} más."
            flash(msg, "warning")
        else:
            flash(msg, "success")
        return redirect(url_for("routes.entradas"))

    except Exception as exc:
        db.session.rollback()
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        flash(f"Error de base de datos: {exc}. No se guardaron cambios.", "danger")
        return redirect(url_for("routes.entradas_importar"))


# ---------------------------------------------------------------------------
# Importar Salidas desde Excel
# ---------------------------------------------------------------------------

@routes_bp.route("/salidas/importar", methods=["GET", "POST"])
@login_required
def salidas_importar():
    import tempfile as _tempfile

    if request.method == "POST":
        if "archivo" not in request.files:
            flash("No se seleccionó ningún archivo.", "danger")
            return redirect(url_for("routes.salidas_importar"))

        file = request.files["archivo"]
        if not file or file.filename == "":
            flash("No se seleccionó ningún archivo.", "danger")
            return redirect(url_for("routes.salidas_importar"))

        if not file.filename.lower().endswith(".xlsx"):
            flash("Solo se aceptan archivos .xlsx (Excel moderno).", "warning")
            return redirect(url_for("routes.salidas_importar"))

        tmp_file = _tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        try:
            file.save(tmp_file.name)
            tmp_path = tmp_file.name
        except Exception:
            tmp_file.close()
            try:
                os.unlink(tmp_file.name)
            except OSError:
                pass
            raise
        finally:
            tmp_file.close()

        # Validar que sea un Excel real (no un archivo con extensión falsa)
        if not _validar_mime_excel(tmp_path):
            flash("El archivo no es un archivo Excel .xlsx válido.", "danger")
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
            return redirect(url_for("routes.salidas_importar"))

        try:
            headers_raw, col_indices, filas, errores_prev = _parse_excel_movimiento("salida", tmp_path)
        except Exception:
            errores_prev = ["Error inesperado al leer el archivo."]
            headers_raw = None
            filas = None

        if errores_prev:
            try:
                os.remove(tmp_path)
            except OSError:
                pass
            for e in errores_prev:
                flash(e, "danger")
            return redirect(url_for("routes.salidas_importar"))

        if not filas:
            try:
                os.remove(tmp_path)
            except OSError:
                pass
            flash("No se encontraron filas válidas para importar.", "warning")
            return redirect(url_for("routes.salidas_importar"))

        headers_mostrar = [h for h in headers_raw if h]
        return render_template(
            "salidas_importar.html",
            preview=True,
            tmp_path=tmp_path,
            headers=headers_mostrar,
            filas=filas[:20],
            total_filas=len(filas),
            nombre_archivo=file.filename,
        )

    return render_template("salidas_importar.html")


@routes_bp.route("/salidas/importar/confirmar", methods=["POST"])
@login_required
def salidas_importar_confirmar():
    tmp_path = request.form.get("tmp_path", "").strip()
    if not tmp_path or not os.path.exists(tmp_path):
        flash("Archivo temporal no encontrado. Vuelve a seleccionar el archivo.", "danger")
        return redirect(url_for("routes.salidas_importar"))

    try:
        headers_raw, col_indices, filas, errores_prev = _parse_excel_movimiento("salida", tmp_path)
    except Exception:
        errores_prev = ["Error inesperado al leer el archivo temporal."]
        filas = None

    if errores_prev:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        for e in errores_prev:
            flash(e, "danger")
        return redirect(url_for("routes.salidas_importar"))

    if not filas:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        flash("No se encontraron filas válidas para importar.", "warning")
        return redirect(url_for("routes.salidas_importar"))

    insertadas = []
    errores = []

    try:
        for f in filas:
            producto = Producto.query.options(joinedload(Producto.familia_rel)).filter_by(codigo=f["codigo"]).first()
            if not producto:
                errores.append(f"'{f['codigo']}': producto no encontrado en BD")
                continue

            if f["cantidad"] > producto.stock_actual:
                errores.append(
                    f"'{f['codigo']}': stock insuficiente "
                    f"(disponible {producto.stock_actual:.2f}, solicitado {f['cantidad']:.2f})"
                )
                continue

            salida = Salida(
                producto_id=producto.id,
                cantidad=f["cantidad"],
                um=f.get("um") or producto.um,
                nro_vale=f.get("nro_vale", ""),
                oi=f.get("oi", ""),
                c_costo=f.get("c_costo", ""),
                maquina=f.get("maquina", ""),
                categoria=f.get("categoria", ""),
                fecha_salida=f.get("fecha") or datetime.now(timezone.utc),
            )
            producto.stock_actual -= f["cantidad"]
            db.session.add(salida)
            insertadas.append(f["codigo"])

        db.session.commit()
        try:
            os.remove(tmp_path)
        except OSError:
            pass

        partes = [f"✅ {len(insertadas)} salidas registradas"]
        if errores:
            partes.append(f"⚠️ {len(errores)} errores")
        msg = "Importación completada: " + ", ".join(partes)
        if errores:
            msg += ". Detalles: " + "; ".join(errores[:5])
            if len(errores) > 5:
                msg += f" y {len(errores) - 5} más."
            flash(msg, "warning")
        else:
            flash(msg, "success")
        return redirect(url_for("routes.salidas"))

    except Exception as exc:
        db.session.rollback()
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        flash(f"Error de base de datos: {exc}. No se guardaron cambios.", "danger")
        return redirect(url_for("routes.salidas_importar"))


# ---------------------------------------------------------------------------
# Ruta: Exportar productos a Excel
# ---------------------------------------------------------------------------

@routes_bp.route("/productos/exportar")
@login_required
def producto_exportar():
    """Exportar todos los productos a un archivo Excel."""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MAESTRA"

    headers = [
        "CODIGO", "COD. CATALOGO", "DESCRIPCION DEL PRODUCTO",
        "U.M", "FAMILIA", "STOCK ACTUAL", "STOCK MINIMO"
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    productos = Producto.query.options(joinedload(Producto.familia_rel)).order_by(Producto.codigo.asc()).all()
    for row_idx, p in enumerate(productos, 2):
        values = [
            p.codigo, p.cod_catalogo, p.descripcion,
            p.um, p.familia, p.stock_actual, p.stock_minimo,
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    widths = [12, 14, 55, 10, 22, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from datetime import date
    today = date.today().strftime("%Y%m%d")

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"productos_{today}.xlsx",
    )

# ---------------------------------------------------------------------------
# Ruta: Reporte maestro combinado
# ---------------------------------------------------------------------------

@routes_bp.route("/reporte-maestro")
@login_required
def reporte_maestro():
    """Página con vista combinada de productos con última entrada y salida."""
    search = request.args.get("search", "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = 25

    query = Producto.query.options(joinedload(Producto.familia_rel))
    if search:
        like = f"%{search}%"
        query = query.filter(
            Producto.codigo.ilike(like)
            | Producto.descripcion.ilike(like)
            | Producto.familia.ilike(like)
        )
    query = query.order_by(Producto.codigo.asc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    # Para cada producto, obtener la última entrada y última salida
    productos_data = []
    for p in pagination.items:
        ultima_entrada = Entrada.query.options(joinedload(Entrada.producto)).filter_by(producto_id=p.id).order_by(Entrada.fecha_ingreso.desc()).first()
        ultima_salida = Salida.query.options(joinedload(Salida.producto)).filter_by(producto_id=p.id).order_by(Salida.fecha_salida.desc()).first()

        productos_data.append({
            "p": p,
            "ultima_entrada": ultima_entrada,
            "ultima_salida": ultima_salida,
        })

    return render_template(
        "reporte_maestro.html",
        productos_data=productos_data,
        pagination=pagination,
        search=search,
    )


def _producto_form(producto=None):
    familias = Familia.query.order_by(Familia.nombre).all()

    if request.method == "POST":
        codigo = request.form.get("codigo", "").strip().upper()
        cod_catalogo = request.form.get("cod_catalogo", "").strip()
        descripcion = request.form.get("descripcion", "").strip()
        um = request.form.get("um", "").strip().upper() or "UND"
        familia = request.form.get("familia", "").strip()
        familia_id = request.form.get("familia_id", "").strip()
        stock_minimo_str = request.form.get("stock_minimo", "0").strip()
        try:
            stock_minimo = float(stock_minimo_str) if stock_minimo_str else 0.0
        except ValueError:
            stock_minimo = 0.0

        errores = []
        if not codigo:
            errores.append("El campo CÓDIGO es obligatorio.")
        if not descripcion:
            errores.append("El campo DESCRIPCIÓN es obligatorio.")

        # Validar código único
        existente = Producto.query.options(joinedload(Producto.familia_rel)).filter(Producto.codigo == codigo).first()
        if existente and (producto is None or existente.id != producto.id):
            errores.append(f"El código '{codigo}' ya está registrado.")

        if errores:
            for e in errores:
                flash(e, "danger")
            return render_template("producto_form.html", producto=producto, valores=request.form, familias=familias)

        is_edit = producto is not None
        if is_edit:
            old_codigo = producto.codigo
            old_descripcion = producto.descripcion
            old_um = producto.um
            old_familia = producto.familia
            old_stock_minimo = producto.stock_minimo

        if producto is None:
            producto = Producto()
            db.session.add(producto)

        producto.codigo = codigo
        producto.cod_catalogo = cod_catalogo
        producto.descripcion = descripcion
        producto.um = um
        producto.familia = familia
        producto.stock_minimo = stock_minimo

        # Set familia_id if selected
        if familia_id:
            try:
                fid = int(familia_id)
                fam = Familia.query.get(fid)
                if fam:
                    producto.familia_id = fid
                    producto.familia = fam.nombre
            except (ValueError, TypeError):
                pass
        else:
            producto.familia_id = None

        if is_edit:
            audit_log("productos", producto.id, "codigo", old_codigo, codigo, current_user.username)
            audit_log("productos", producto.id, "descripcion", old_descripcion, descripcion, current_user.username)
            audit_log("productos", producto.id, "um", old_um, um, current_user.username)
            audit_log("productos", producto.id, "familia", old_familia, familia, current_user.username)
            audit_log("productos", producto.id, "stock_minimo", old_stock_minimo, stock_minimo, current_user.username)

        db.session.commit()
        flash("Producto guardado correctamente.", "success")
        return redirect(url_for("routes.productos"))

    return render_template("producto_form.html", producto=producto, familias=familias)


# ---------------------------------------------------------------------------
# Registrar Entrada
# ---------------------------------------------------------------------------

@routes_bp.route("/entradas", methods=["GET", "POST"])
@login_required
def entradas():
    productos_list = Producto.query.options(joinedload(Producto.familia_rel)).order_by(Producto.descripcion.asc()).all()
    ordenes_compra = OrdenCompra.query.order_by(OrdenCompra.numero.asc()).all()

    if request.method == "POST":
        producto_id = request.form.get("producto_id")
        cantidad_str = request.form.get("cantidad", "0").strip()
        zona = request.form.get("zona", "").strip()
        ubicacion = request.form.get("ubicacion", "").strip()
        alm = request.form.get("alm", "ALM-01").strip()
        oc = request.form.get("oc", "").strip()
        guia_remision = request.form.get("guia_remision", "").strip()
        familia = request.form.get("familia", "").strip()
        fecha_str = request.form.get("fecha_ingreso", "").strip()

        errores = []
        if not producto_id:
            errores.append("Debe seleccionar un producto.")
        try:
            cantidad = float(cantidad_str) if cantidad_str else 0.0
        except ValueError:
            cantidad = 0.0
            errores.append("La cantidad debe ser un número válido.")
        if cantidad <= 0:
            errores.append("La cantidad debe ser mayor a 0.")

        # Parsear fecha
        fecha_ingreso = datetime.now(timezone.utc)
        if fecha_str:
            try:
                fecha_ingreso = datetime.strptime(fecha_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            except ValueError:
                errores.append("Formato de fecha inválido (use YYYY-MM-DD).")

        if errores:
            for e in errores:
                flash(e, "danger")
            return render_template("entrada_form.html", productos=productos_list, valores=request.form, ordenes_compra=ordenes_compra)

        producto = db.session.get(Producto, int(producto_id))
        if not producto:
            flash("Producto no encontrado.", "danger")
            return redirect(url_for("routes.entradas"))

        entrada = Entrada(
            producto_id=producto.id,
            cantidad=cantidad,
            um=producto.um,
            zona=zona,
            ubicacion=ubicacion,
            alm=alm,
            fecha_ingreso=fecha_ingreso,
            oc=oc,
            guia_remision=guia_remision,
            familia=familia or producto.familia,
        )
        # Vincular OC por número si existe
        if oc:
            oc_obj = OrdenCompra.query.filter_by(numero=oc).first()
            if oc_obj:
                entrada.oc_id = oc_obj.id

        producto.stock_actual += cantidad

        db.session.add(entrada)
        db.session.commit()
        flash(f"Entrada registrada. Stock actualizado: {producto.descripcion} → {producto.stock_actual:.2f} {producto.um}", "success")
        return redirect(url_for("routes.entradas"))

    return render_template("entrada_form.html", productos=productos_list, ordenes_compra=ordenes_compra)


@routes_bp.route("/entradas/editar/<int:entrada_id>", methods=["GET", "POST"])
@login_required
def entrada_editar(entrada_id):
    """Editar una entrada existente. Ajusta el stock del producto."""
    entrada = db.session.get(Entrada, entrada_id)
    if not entrada:
        flash("Entrada no encontrada.", "danger")
        return redirect(url_for("routes.historial"))
    producto = entrada.producto
    productos_list = Producto.query.options(joinedload(Producto.familia_rel)).order_by(Producto.descripcion.asc()).all()
    ordenes_compra = OrdenCompra.query.order_by(OrdenCompra.numero.asc()).all()

    if request.method == "POST":
        cantidad_str = request.form.get("cantidad", "0").strip()
        zona = request.form.get("zona", "").strip()
        ubicacion = request.form.get("ubicacion", "").strip()
        alm = request.form.get("alm", "").strip()
        oc = request.form.get("oc", "").strip()
        guia_remision = request.form.get("guia_remision", "").strip()
        familia = request.form.get("familia", "").strip()
        fecha_str = request.form.get("fecha_ingreso", "").strip()
        oc_id_str = request.form.get("oc_id", "").strip()

        errores = []
        try:
            nueva_cantidad = float(cantidad_str) if cantidad_str else 0.0
        except ValueError:
            nueva_cantidad = 0.0
            errores.append("La cantidad debe ser un número válido.")
        if nueva_cantidad <= 0:
            errores.append("La cantidad debe ser mayor a 0.")

        fecha_ingreso = entrada.fecha_ingreso
        if fecha_str:
            try:
                fecha_ingreso = datetime.strptime(fecha_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            except ValueError:
                errores.append("Formato de fecha inválido (use YYYY-MM-DD).")

        if errores:
            for e in errores:
                flash(e, "danger")
            return render_template("entrada_form.html", entrada=entrada, productos=productos_list, valores=request.form, ordenes_compra=ordenes_compra)

        # Capturar valores anteriores para auditoría
        old_cantidad = entrada.cantidad
        old_zona = entrada.zona
        old_ubicacion = entrada.ubicacion
        old_alm = entrada.alm
        old_oc = entrada.oc
        old_guia_remision = entrada.guia_remision
        old_fecha_ingreso = entrada.fecha_ingreso

        # Ajustar stock: revertir cantidad anterior y aplicar nueva
        diferencia = nueva_cantidad - entrada.cantidad
        producto.stock_actual += diferencia

        entrada.cantidad = nueva_cantidad
        entrada.zona = zona
        entrada.ubicacion = ubicacion
        entrada.alm = alm or entrada.alm
        entrada.oc = oc
        entrada.guia_remision = guia_remision
        entrada.familia = familia or entrada.familia
        entrada.fecha_ingreso = fecha_ingreso
        entrada.um = producto.um

        # Actualizar vínculo con OC
        if oc_id_str:
            try:
                oc_id_val = int(oc_id_str)
                if oc_id_val == 0:
                    entrada.oc_id = None
                else:
                    oc_obj = db.session.get(OrdenCompra, oc_id_val)
                    if oc_obj:
                        entrada.oc_id = oc_id_val
            except (ValueError, TypeError):
                pass
        else:
            # Auto-link by number if oc_id not specified
            if oc:
                oc_obj = OrdenCompra.query.filter_by(numero=oc).first()
                if oc_obj:
                    entrada.oc_id = oc_obj.id
                else:
                    entrada.oc_id = None
            else:
                entrada.oc_id = None

        audit_log("entradas", entrada.id, "cantidad", old_cantidad, nueva_cantidad, current_user.username)
        audit_log("entradas", entrada.id, "zona", old_zona, zona, current_user.username)
        audit_log("entradas", entrada.id, "ubicacion", old_ubicacion, ubicacion, current_user.username)
        audit_log("entradas", entrada.id, "alm", old_alm, alm or old_alm, current_user.username)
        audit_log("entradas", entrada.id, "oc", old_oc, oc, current_user.username)
        audit_log("entradas", entrada.id, "guia_remision", old_guia_remision, guia_remision, current_user.username)
        audit_log("entradas", entrada.id, "fecha_ingreso", str(old_fecha_ingreso), str(fecha_ingreso), current_user.username)

        db.session.commit()
        flash(f"Entrada actualizada. Stock de {producto.descripcion}: {producto.stock_actual:.2f} {producto.um}", "success")
        return redirect(url_for("routes.historial"))

    return render_template("entrada_form.html", entrada=entrada, productos=productos_list, ordenes_compra=ordenes_compra)


@routes_bp.route("/entradas/eliminar/<int:entrada_id>", methods=["POST"])
@login_required
def entrada_eliminar(entrada_id):
    """Eliminar una entrada. Revierte el stock del producto."""
    entrada = db.session.get(Entrada, entrada_id)
    if not entrada:
        flash("Entrada no encontrada.", "danger")
        return redirect(url_for("routes.historial"))

    producto = entrada.producto
    producto.stock_actual -= entrada.cantidad
    desc = producto.descripcion
    db.session.delete(entrada)
    db.session.commit()
    flash(f"Entrada eliminada. Stock de {desc}: {producto.stock_actual:.2f} {producto.um}", "success")
    return redirect(url_for("routes.historial"))


# ---------------------------------------------------------------------------
# Registrar Salida
# ---------------------------------------------------------------------------

@routes_bp.route("/salidas", methods=["GET", "POST"])
@login_required
def salidas():
    productos_list = Producto.query.options(joinedload(Producto.familia_rel)).order_by(Producto.descripcion.asc()).all()

    if request.method == "POST":
        producto_id = request.form.get("producto_id")
        cantidad_str = request.form.get("cantidad", "0").strip()
        nro_vale = request.form.get("nro_vale", "").strip()
        oi = request.form.get("oi", "").strip()
        c_costo = request.form.get("c_costo", "").strip()
        maquina = request.form.get("maquina", "").strip()
        categoria = request.form.get("categoria", "").strip()
        fecha_str = request.form.get("fecha_salida", "").strip()

        errores = []
        if not producto_id:
            errores.append("Debe seleccionar un producto.")
        try:
            cantidad = float(cantidad_str) if cantidad_str else 0.0
        except ValueError:
            cantidad = 0.0
            errores.append("La cantidad debe ser un número válido.")
        if cantidad <= 0:
            errores.append("La cantidad debe ser mayor a 0.")

        fecha_salida = datetime.now(timezone.utc)
        if fecha_str:
            try:
                fecha_salida = datetime.strptime(fecha_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            except ValueError:
                errores.append("Formato de fecha inválido (use YYYY-MM-DD).")

        if errores:
            for e in errores:
                flash(e, "danger")
            return render_template("salida_form.html", productos=productos_list, valores=request.form)

        producto = db.session.get(Producto, int(producto_id))
        if not producto:
            flash("Producto no encontrado.", "danger")
            return redirect(url_for("routes.salidas"))

        if cantidad > producto.stock_actual:
            flash(f"Stock insuficiente. Disponible: {producto.stock_actual:.2f} {producto.um}", "danger")
            return render_template("salida_form.html", productos=productos_list, valores=request.form)

        salida = Salida(
            producto_id=producto.id,
            cantidad=cantidad,
            um=producto.um,
            fecha_salida=fecha_salida,
            nro_vale=nro_vale,
            oi=oi,
            c_costo=c_costo,
            maquina=maquina,
            categoria=categoria,
        )
        producto.stock_actual -= cantidad

        db.session.add(salida)
        db.session.commit()
        flash(f"Salida registrada. Stock actualizado: {producto.descripcion} → {producto.stock_actual:.2f} {producto.um}", "success")
        return redirect(url_for("routes.salidas"))

    return render_template("salida_form.html", productos=productos_list)


@routes_bp.route("/salidas/editar/<int:salida_id>", methods=["GET", "POST"])
@login_required
def salida_editar(salida_id):
    """Editar una salida existente. Ajusta el stock del producto."""
    salida = db.session.get(Salida, salida_id)
    if not salida:
        flash("Salida no encontrada.", "danger")
        return redirect(url_for("routes.historial"))
    producto = salida.producto
    productos_list = Producto.query.options(joinedload(Producto.familia_rel)).order_by(Producto.descripcion.asc()).all()

    if request.method == "POST":
        cantidad_str = request.form.get("cantidad", "0").strip()
        nro_vale = request.form.get("nro_vale", "").strip()
        oi = request.form.get("oi", "").strip()
        c_costo = request.form.get("c_costo", "").strip()
        maquina = request.form.get("maquina", "").strip()
        categoria = request.form.get("categoria", "").strip()
        fecha_str = request.form.get("fecha_salida", "").strip()

        errores = []
        try:
            nueva_cantidad = float(cantidad_str) if cantidad_str else 0.0
        except ValueError:
            nueva_cantidad = 0.0
            errores.append("La cantidad debe ser un número válido.")
        if nueva_cantidad <= 0:
            errores.append("La cantidad debe ser mayor a 0.")

        fecha_salida = salida.fecha_salida
        if fecha_str:
            try:
                fecha_salida = datetime.strptime(fecha_str, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            except ValueError:
                errores.append("Formato de fecha inválido (use YYYY-MM-DD).")

        if errores:
            for e in errores:
                flash(e, "danger")
            return render_template("salida_form.html", salida=salida, productos=productos_list, valores=request.form)

        # Capturar valores anteriores para auditoría
        old_cantidad = salida.cantidad
        old_nro_vale = salida.nro_vale
        old_oi = salida.oi
        old_c_costo = salida.c_costo
        old_maquina = salida.maquina
        old_categoria = salida.categoria
        old_fecha_salida = salida.fecha_salida

        # Calcular diferencia y validar stock disponible
        diferencia = nueva_cantidad - salida.cantidad
        if diferencia > 0 and diferencia > producto.stock_actual:
            flash(
                f"Stock insuficiente. Disponible: {producto.stock_actual:.2f} {producto.um}, "
                f"incremento solicitado: {diferencia:.2f}",
                "danger",
            )
            return render_template("salida_form.html", salida=salida, productos=productos_list, valores=request.form)

        producto.stock_actual -= diferencia
        salida.cantidad = nueva_cantidad
        salida.nro_vale = nro_vale
        salida.oi = oi
        salida.c_costo = c_costo
        salida.maquina = maquina
        salida.categoria = categoria
        salida.fecha_salida = fecha_salida
        salida.um = producto.um

        audit_log("salidas", salida.id, "cantidad", old_cantidad, nueva_cantidad, current_user.username)
        audit_log("salidas", salida.id, "nro_vale", old_nro_vale, nro_vale, current_user.username)
        audit_log("salidas", salida.id, "oi", old_oi, oi, current_user.username)
        audit_log("salidas", salida.id, "c_costo", old_c_costo, c_costo, current_user.username)
        audit_log("salidas", salida.id, "maquina", old_maquina, maquina, current_user.username)
        audit_log("salidas", salida.id, "categoria", old_categoria, categoria, current_user.username)
        audit_log("salidas", salida.id, "fecha_salida", str(old_fecha_salida), str(fecha_salida), current_user.username)

        db.session.commit()
        flash(f"Salida actualizada. Stock de {producto.descripcion}: {producto.stock_actual:.2f} {producto.um}", "success")
        return redirect(url_for("routes.historial"))

    return render_template("salida_form.html", salida=salida, productos=productos_list)


@routes_bp.route("/salidas/eliminar/<int:salida_id>", methods=["POST"])
@login_required
def salida_eliminar(salida_id):
    """Eliminar una salida. Revierte el stock del producto."""
    salida = db.session.get(Salida, salida_id)
    if not salida:
        flash("Salida no encontrada.", "danger")
        return redirect(url_for("routes.historial"))

    producto = salida.producto
    producto.stock_actual += salida.cantidad
    desc = producto.descripcion
    db.session.delete(salida)
    db.session.commit()
    flash(f"Salida eliminada. Stock de {desc}: {producto.stock_actual:.2f} {producto.um}", "success")
    return redirect(url_for("routes.historial"))


# ---------------------------------------------------------------------------
# Consulta de Existencias
# ---------------------------------------------------------------------------

@routes_bp.route("/existencias")
@login_required
def existencias():
    search = request.args.get("search", "").strip()
    familia = request.args.get("familia", "").strip()
    solo_bajo = request.args.get("solo_bajo", "").strip()
    con_saldo = request.args.get("con_saldo", "1").strip()  # default: solo con saldo
    page = request.args.get("page", 1, type=int)
    per_page = 20

    query = Producto.query.options(joinedload(Producto.familia_rel))
    if search:
        query = query.filter(
            Producto.descripcion.ilike(f"%{search}%")
            | Producto.codigo.ilike(f"%{search}%")
        )
    if familia:
        query = query.filter(Producto.familia.ilike(f"%{familia}%"))
    if solo_bajo == "1":
        query = query.filter(
            Producto.stock_minimo > 0,
            Producto.stock_actual <= Producto.stock_minimo
        )
    if con_saldo == "1":
        query = query.filter(Producto.stock_actual > 0)

    query = query.order_by(Producto.descripcion.asc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    # Obtener lista de familias para el filtro
    familias = Familia.query.order_by(Familia.nombre).all()

    return render_template(
        "existencias.html",
        pagination=pagination,
        search=search,
        familia=familia,
        solo_bajo=solo_bajo,
        con_saldo=con_saldo,
        familias=familias,
    )


@routes_bp.route("/existencias/exportar")
@login_required
def existencias_exportar():
    """Exportar existencias filtradas a Excel."""

    search = request.args.get("search", "").strip()
    familia = request.args.get("familia", "").strip()
    solo_bajo = request.args.get("solo_bajo", "").strip()
    con_saldo = request.args.get("con_saldo", "1").strip()

    query = Producto.query.options(joinedload(Producto.familia_rel))
    if search:
        query = query.filter(
            Producto.descripcion.ilike(f"%{search}%")
            | Producto.codigo.ilike(f"%{search}%")
        )
    if familia:
        query = query.filter(Producto.familia.ilike(f"%{familia}%"))
    if solo_bajo == "1":
        query = query.filter(
            Producto.stock_minimo > 0,
            Producto.stock_actual <= Producto.stock_minimo
        )
    if con_saldo == "1":
        query = query.filter(Producto.stock_actual > 0)
    productos = query.order_by(Producto.descripcion.asc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "EXISTENCIAS"

    headers = ["CÓDIGO", "COD. CATÁLOGO", "DESCRIPCIÓN", "U.M",
               "FAMILIA", "STOCK ACTUAL", "STOCK MÍNIMO", "ESTADO"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = header_font; c.fill = header_fill; c.border = thin_border

    for ri, p in enumerate(productos, 2):
        estado = "STOCK BAJO" if (p.stock_bajo and p.stock_minimo > 0) else \
                 "SIN STOCK" if p.stock_actual == 0 else "OK"
        vals = [p.codigo, p.cod_catalogo, p.descripcion, p.um,
                p.familia, p.stock_actual, p.stock_minimo, estado]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = thin_border

    widths = [12, 14, 55, 10, 22, 14, 14, 14]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    from datetime import date
    return send_file(output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=f"existencias_{date.today().strftime('%Y%m%d')}.xlsx")


# ---------------------------------------------------------------------------
# Historial de Movimientos
# ---------------------------------------------------------------------------

@routes_bp.route("/historial")
@login_required
def historial():
    producto_search = request.args.get("producto", "").strip()
    fecha_desde = request.args.get("fecha_desde", "").strip()
    fecha_hasta = request.args.get("fecha_hasta", "").strip()
    tipo = request.args.get("tipo", "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = 20

    # Construir lista combinada
    movimientos = []

    # Entradas
    entradas_query = Entrada.query.options(joinedload(Entrada.producto))
    if producto_search:
        entradas_query = entradas_query.join(Producto).filter(
            Producto.descripcion.ilike(f"%{producto_search}%")
            | Producto.codigo.ilike(f"%{producto_search}%")
        )
    if fecha_desde:
        try:
            fd = datetime.strptime(fecha_desde, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            entradas_query = entradas_query.filter(Entrada.fecha_ingreso >= fd)
        except ValueError:
            pass
    if fecha_hasta:
        try:
            fh = datetime.strptime(fecha_hasta, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            entradas_query = entradas_query.filter(Entrada.fecha_ingreso <= fh)
        except ValueError:
            pass

    salidas_query = Salida.query.options(joinedload(Salida.producto))
    if producto_search:
        salidas_query = salidas_query.join(Producto).filter(
            Producto.descripcion.ilike(f"%{producto_search}%")
            | Producto.codigo.ilike(f"%{producto_search}%")
        )
    if fecha_desde:
        try:
            fd = datetime.strptime(fecha_desde, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            salidas_query = salidas_query.filter(Salida.fecha_salida >= fd)
        except ValueError:
            pass
    if fecha_hasta:
        try:
            fh = datetime.strptime(fecha_hasta, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            salidas_query = salidas_query.filter(Salida.fecha_salida <= fh)
        except ValueError:
            pass

    # Paginación directa en BD con UNION ALL
    from sqlalchemy import literal_column, union_all

    if tipo == "SALIDA":
        pag = salidas_query.order_by(Salida.fecha_salida.desc()).paginate(page=page, per_page=per_page, error_out=False)
        movimientos_paginados = []
        for s in pag.items:
            if s.producto:
                movimientos_paginados.append({
                    "id": f"S-{s.id}",
                    "tipo": "SALIDA",
                    "fecha": s.fecha_salida,
                    "producto": s.producto.descripcion,
                    "codigo": s.producto.codigo,
                    "cantidad": s.cantidad,
                    "um": s.um,
                    "referencia": s.nro_vale or s.oi or "—",
                    "detalle": f"Vale: {s.nro_vale}, OI: {s.oi}, C.Costo: {s.c_costo}, Máq: {s.maquina}",
                })
        total = pag.total
    elif tipo == "ENTRADA":
        pag = entradas_query.order_by(Entrada.fecha_ingreso.desc()).paginate(page=page, per_page=per_page, error_out=False)
        movimientos_paginados = []
        for e in pag.items:
            if e.producto:
                movimientos_paginados.append({
                    "id": f"E-{e.id}",
                    "tipo": "ENTRADA",
                    "fecha": e.fecha_ingreso,
                    "producto": e.producto.descripcion,
                    "codigo": e.producto.codigo,
                    "cantidad": e.cantidad,
                    "um": e.um,
                    "referencia": e.oc or e.guia_remision or "—",
                    "detalle": f"OC: {e.oc}, Guía: {e.guia_remision}, Zona: {e.zona}, Ubic: {e.ubicacion}",
                })
        total = pag.total
    else:
        # UNION ALL para paginación combinada
        from sqlalchemy import text as sa_text

        e_q = entradas_query.with_entities(
            literal_column("'E'").label('src'),
            Entrada.id.label('id'),
            Entrada.fecha_ingreso.label('fecha')
        )
        s_q = salidas_query.with_entities(
            literal_column("'S'").label('src'),
            Salida.id.label('id'),
            Salida.fecha_salida.label('fecha')
        )

        union_q = e_q.union_all(s_q)

        # Total combinado
        total = db.session.query(db.func.count()).select_from(union_q.subquery()).scalar() or 0

        # Paginación ordenada por fecha
        paginated = union_q.order_by(sa_text('fecha DESC')).limit(per_page).offset((page - 1) * per_page).all()

        movimientos_paginados = []
        for src, mov_id, _ in paginated:
            if src == 'E':
                e = db.session.get(Entrada, mov_id)
                if e and e.producto:
                    movimientos_paginados.append({
                        "id": f"E-{e.id}",
                        "tipo": "ENTRADA",
                        "fecha": e.fecha_ingreso,
                        "producto": e.producto.descripcion,
                        "codigo": e.producto.codigo,
                        "cantidad": e.cantidad,
                        "um": e.um,
                        "referencia": e.oc or e.guia_remision or "—",
                        "detalle": f"OC: {e.oc}, Guía: {e.guia_remision}, Zona: {e.zona}, Ubic: {e.ubicacion}",
                    })
            else:
                s = db.session.get(Salida, mov_id)
                if s and s.producto:
                    movimientos_paginados.append({
                        "id": f"S-{s.id}",
                        "tipo": "SALIDA",
                        "fecha": s.fecha_salida,
                        "producto": s.producto.descripcion,
                        "codigo": s.producto.codigo,
                        "cantidad": s.cantidad,
                        "um": s.um,
                        "referencia": s.nro_vale or s.oi or "—",
                        "detalle": f"Vale: {s.nro_vale}, OI: {s.oi}, C.Costo: {s.c_costo}, Máq: {s.maquina}",
                    })

    # Objeto de paginación simple
    class SimplePagination:
        def __init__(self, items, total, page, per_page):
            self.items = items
            self.total = total
            self.page = page
            self.per_page = per_page
            self.pages = max(1, (total + per_page - 1) // per_page)

        @property
        def has_prev(self):
            return self.page > 1

        @property
        def has_next(self):
            return self.page < self.pages

        @property
        def prev_num(self):
            return self.page - 1

        @property
        def next_num(self):
            return self.page + 1

        def iter_pages(self, left_edge=1, left_current=2, right_current=3, right_edge=1):
            last = 0
            for num in range(1, self.pages + 1):
                if (
                    num <= left_edge
                    or (num > self.page - left_current - 1 and num < self.page + right_current)
                    or num > self.pages - right_edge
                ):
                    if last + 1 != num:
                        yield None
                    yield num
                    last = num

    pagination = SimplePagination(movimientos_paginados, total, page, per_page)

    return render_template(
        "historial.html",
        pagination=pagination,
        producto=producto_search,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        tipo=tipo,
    )


@routes_bp.route("/historial/exportar")
@login_required
def historial_exportar():
    """Exportar historial de movimientos filtrado a Excel."""

    producto_search = request.args.get("producto", "").strip()
    fecha_desde = request.args.get("fecha_desde", "").strip()
    fecha_hasta = request.args.get("fecha_hasta", "").strip()
    tipo = request.args.get("tipo", "").strip()

    # Construir datos igual que en historial() pero sin paginar
    movimientos = []

    entradas_query = Entrada.query
    salidas_query = Salida.query

    if producto_search:
        entradas_query = entradas_query.join(Producto).filter(
            Producto.descripcion.ilike(f"%{producto_search}%")
            | Producto.codigo.ilike(f"%{producto_search}%")
        )
        salidas_query = salidas_query.join(Producto).filter(
            Producto.descripcion.ilike(f"%{producto_search}%")
            | Producto.codigo.ilike(f"%{producto_search}%")
        )
    if fecha_desde:
        try:
            fd = datetime.strptime(fecha_desde, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            entradas_query = entradas_query.filter(Entrada.fecha_ingreso >= fd)
            salidas_query = salidas_query.filter(Salida.fecha_salida >= fd)
        except ValueError:
            pass
    if fecha_hasta:
        try:
            fh = datetime.strptime(fecha_hasta, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            entradas_query = entradas_query.filter(Entrada.fecha_ingreso <= fh)
            salidas_query = salidas_query.filter(Salida.fecha_salida <= fh)
        except ValueError:
            pass

    if tipo != "SALIDA":
        for e in entradas_query.options(joinedload(Entrada.producto)).order_by(Entrada.fecha_ingreso.desc()).all():
            movimientos.append([
                e.fecha_ingreso.strftime("%Y-%m-%d %H:%M") if e.fecha_ingreso else "",
                "ENTRADA",
                e.producto.codigo if e.producto else "",
                e.producto.descripcion if e.producto else "",
                e.cantidad, e.um or "",
                e.oc or "", e.guia_remision or "", e.zona or "",
                e.ubicacion or "", e.alm or "", e.familia or "",
            ])
    if tipo != "ENTRADA":
        for s in salidas_query.options(joinedload(Salida.producto)).order_by(Salida.fecha_salida.desc()).all():
            movimientos.append([
                s.fecha_salida.strftime("%Y-%m-%d %H:%M") if s.fecha_salida else "",
                "SALIDA",
                s.producto.codigo if s.producto else "",
                s.producto.descripcion if s.producto else "",
                s.cantidad, s.um or "",
                s.nro_vale or "", s.oi or "", s.c_costo or "",
                s.maquina or "", s.categoria or "", "",
            ])

    movimientos.sort(key=lambda r: r[0], reverse=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MOVIMIENTOS"

    headers = ["FECHA", "TIPO", "CÓDIGO", "PRODUCTO", "CANTIDAD", "U.M",
               "OC/VALE", "GUÍA/OI", "ZONA/C.COSTO", "UBIC/MÁQ", "ALM/CAT", "FAMILIA"]
    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    tb = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hf; c.fill = hfill; c.border = tb

    for ri, row in enumerate(movimientos, 2):
        for ci, v in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = tb

    widths = [18, 10, 12, 50, 10, 8, 14, 14, 14, 14, 14, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output); output.seek(0)
    from datetime import date
    return send_file(output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True, download_name=f"historial_{date.today().strftime('%Y%m%d')}.xlsx")


# ===========================================================================
# CRUD Familias (solo administrador)
# ===========================================================================

@routes_bp.route("/familias")
@login_required
def familias():
    """Listar todas las familias con paginación."""
    check = _admin_required()
    if check:
        return check
    page = request.args.get("page", 1, type=int)
    per_page = 20
    pagination = Familia.query.order_by(Familia.nombre).paginate(page=page, per_page=per_page, error_out=False)
    return render_template("familias.html", pagination=pagination)


@routes_bp.route("/familias/nuevo", methods=["GET", "POST"])
@login_required
def familia_nuevo():
    """Crear nueva familia."""
    check = _admin_required()
    if check:
        return check

    if request.method == "POST":
        nombre = request.form.get("nombre", "").strip().upper()
        color = request.form.get("color", "#6c757d").strip()

        errores = []
        if not nombre:
            errores.append("El nombre de la familia es obligatorio.")
        if Familia.query.filter_by(nombre=nombre).first():
            errores.append(f"La familia '{nombre}' ya existe.")

        if errores:
            for e in errores:
                flash(e, "danger")
            return render_template("familia_form.html", valores=request.form)

        familia = Familia(nombre=nombre, color=color)
        db.session.add(familia)
        db.session.commit()
        flash(f"Familia '{nombre}' creada correctamente.", "success")
        return redirect(url_for("routes.familias"))

    return render_template("familia_form.html")


@routes_bp.route("/familias/editar/<int:familia_id>", methods=["GET", "POST"])
@login_required
def familia_editar(familia_id):
    """Editar familia existente."""
    check = _admin_required()
    if check:
        return check

    familia = db.session.get(Familia, familia_id)
    if not familia:
        flash("Familia no encontrada.", "danger")
        return redirect(url_for("routes.familias"))

    if request.method == "POST":
        nombre = request.form.get("nombre", "").strip().upper()
        color = request.form.get("color", "#6c757d").strip()

        errores = []
        if not nombre:
            errores.append("El nombre de la familia es obligatorio.")

        # Validar nombre único (excluyendo la familia actual)
        existente = Familia.query.filter(Familia.nombre == nombre, Familia.id != familia_id).first()
        if existente:
            errores.append(f"La familia '{nombre}' ya existe.")

        if errores:
            for e in errores:
                flash(e, "danger")
            return render_template("familia_form.html", familia=familia, valores=request.form)

        old_nombre = familia.nombre
        old_color = familia.color
        familia.nombre = nombre
        familia.color = color

        audit_log("familias", familia.id, "nombre", old_nombre, nombre, current_user.username)
        audit_log("familias", familia.id, "color", old_color, color, current_user.username)

        db.session.commit()
        flash(f"Familia '{nombre}' actualizada correctamente.", "success")
        return redirect(url_for("routes.familias"))

    return render_template("familia_form.html", familia=familia)


@routes_bp.route("/familias/eliminar/<int:familia_id>", methods=["POST"])
@login_required
def familia_eliminar(familia_id):
    """Eliminar familia solo si no tiene productos asociados."""
    check = _admin_required()
    if check:
        return check

    familia = db.session.get(Familia, familia_id)
    if not familia:
        flash("Familia no encontrada.", "danger")
        return redirect(url_for("routes.familias"))

    if familia.productos.count() > 0:
        flash(f"No se puede eliminar la familia '{familia.nombre}' porque tiene {familia.productos.count()} producto(s) asociado(s).", "danger")
        return redirect(url_for("routes.familias"))

    db.session.delete(familia)
    db.session.commit()
    flash(f"Familia '{familia.nombre}' eliminada correctamente.", "success")
    return redirect(url_for("routes.familias"))


# ===========================================================================
# CRUD Almacenes (solo administrador)
# ===========================================================================

@routes_bp.route("/almacenes")
@login_required
def almacenes():
    """Listar todos los almacenes con paginación."""
    check = _admin_required()
    if check:
        return check
    page = request.args.get("page", 1, type=int)
    per_page = 20
    pagination = Almacen.query.order_by(Almacen.codigo).paginate(page=page, per_page=per_page, error_out=False)
    return render_template("almacenes.html", pagination=pagination)


@routes_bp.route("/almacenes/nuevo", methods=["GET", "POST"])
@login_required
def almacen_nuevo():
    """Crear nuevo almacén."""
    check = _admin_required()
    if check:
        return check

    if request.method == "POST":
        codigo = request.form.get("codigo", "").strip().upper()
        nombre = request.form.get("nombre", "").strip()
        direccion = request.form.get("direccion", "").strip()

        errores = []
        if not codigo:
            errores.append("El código del almacén es obligatorio.")
        if not nombre:
            errores.append("El nombre del almacén es obligatorio.")
        if Almacen.query.filter_by(codigo=codigo).first():
            errores.append(f"El código '{codigo}' ya está registrado.")

        if errores:
            for e in errores:
                flash(e, "danger")
            return render_template("almacen_form.html", valores=request.form)

        almacen = Almacen(codigo=codigo, nombre=nombre, direccion=direccion)
        db.session.add(almacen)
        db.session.commit()
        flash(f"Almacén '{nombre}' creado correctamente.", "success")
        return redirect(url_for("routes.almacenes"))

    return render_template("almacen_form.html")


@routes_bp.route("/almacenes/editar/<int:almacen_id>", methods=["GET", "POST"])
@login_required
def almacen_editar(almacen_id):
    """Editar almacén existente."""
    check = _admin_required()
    if check:
        return check

    almacen = db.session.get(Almacen, almacen_id)
    if not almacen:
        flash("Almacén no encontrado.", "danger")
        return redirect(url_for("routes.almacenes"))

    if request.method == "POST":
        codigo = request.form.get("codigo", "").strip().upper()
        nombre = request.form.get("nombre", "").strip()
        direccion = request.form.get("direccion", "").strip()

        errores = []
        if not codigo:
            errores.append("El código del almacén es obligatorio.")
        if not nombre:
            errores.append("El nombre del almacén es obligatorio.")

        # Validar código único (excluyendo el almacén actual)
        existente = Almacen.query.filter(Almacen.codigo == codigo, Almacen.id != almacen_id).first()
        if existente:
            errores.append(f"El código '{codigo}' ya está registrado.")

        if errores:
            for e in errores:
                flash(e, "danger")
            return render_template("almacen_form.html", almacen=almacen, valores=request.form)

        old_codigo = almacen.codigo
        old_nombre = almacen.nombre
        old_direccion = almacen.direccion

        almacen.codigo = codigo
        almacen.nombre = nombre
        almacen.direccion = direccion

        audit_log("almacenes", almacen.id, "codigo", old_codigo, codigo, current_user.username)
        audit_log("almacenes", almacen.id, "nombre", old_nombre, nombre, current_user.username)
        audit_log("almacenes", almacen.id, "direccion", old_direccion, direccion, current_user.username)

        db.session.commit()
        flash(f"Almacén '{nombre}' actualizado correctamente.", "success")
        return redirect(url_for("routes.almacenes"))

    return render_template("almacen_form.html", almacen=almacen)


@routes_bp.route("/almacenes/eliminar/<int:almacen_id>", methods=["POST"])
@login_required
def almacen_eliminar(almacen_id):
    """Eliminar almacén solo si no tiene productos asociados."""
    check = _admin_required()
    if check:
        return check

    almacen = db.session.get(Almacen, almacen_id)
    if not almacen:
        flash("Almacén no encontrado.", "danger")
        return redirect(url_for("routes.almacenes"))

    if almacen.productos.count() > 0:
        flash(f"No se puede eliminar el almacén '{almacen.nombre}' porque tiene {almacen.productos.count()} producto(s) asociado(s).", "danger")
        return redirect(url_for("routes.almacenes"))

    db.session.delete(almacen)
    db.session.commit()
    flash(f"Almacén '{almacen.nombre}' eliminado correctamente.", "success")
    return redirect(url_for("routes.almacenes"))


# ===========================================================================
# CRUD Órdenes de Compra
# ===========================================================================

@routes_bp.route("/oc")
@login_required
def oc_lista():
    """Listar órdenes de compra con paginación y filtros."""
    search = request.args.get("search", "").strip()
    estado_filtro = request.args.get("estado", "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = 20

    query = OrdenCompra.query
    if search:
        like = f"%{search}%"
        query = query.filter(
            OrdenCompra.numero.ilike(like)
            | OrdenCompra.proveedor.ilike(like)
        )
    if estado_filtro:
        query = query.filter(OrdenCompra.estado == estado_filtro.upper())

    query = query.order_by(OrdenCompra.created_at.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    return render_template("oc_lista.html", pagination=pagination, search=search, estado_filtro=estado_filtro)


@routes_bp.route("/oc/nuevo", methods=["GET", "POST"])
@login_required
def oc_nuevo():
    """Crear nueva orden de compra."""
    if request.method == "POST":
        numero = request.form.get("numero", "").strip().upper()
        proveedor = request.form.get("proveedor", "").strip()
        estado = request.form.get("estado", "PENDIENTE").strip().upper()

        errores = []
        if not numero:
            errores.append("El número de orden de compra es obligatorio.")
        if OrdenCompra.query.filter_by(numero=numero).first():
            errores.append(f"La OC '{numero}' ya existe.")

        if errores:
            for e in errores:
                flash(e, "danger")
            return render_template("oc_form.html", valores=request.form)

        oc = OrdenCompra(numero=numero, proveedor=proveedor, estado=estado)
        db.session.add(oc)
        db.session.commit()
        flash(f"OC '{numero}' creada correctamente.", "success")
        return redirect(url_for("routes.oc_lista"))

    return render_template("oc_form.html")


@routes_bp.route("/oc/editar/<int:oc_id>", methods=["GET", "POST"])
@login_required
def oc_editar(oc_id):
    """Editar orden de compra existente."""
    oc = db.session.get(OrdenCompra, oc_id)
    if not oc:
        flash("Orden de compra no encontrada.", "danger")
        return redirect(url_for("routes.oc_lista"))

    if request.method == "POST":
        numero = request.form.get("numero", "").strip().upper()
        proveedor = request.form.get("proveedor", "").strip()
        estado = request.form.get("estado", "PENDIENTE").strip().upper()

        errores = []
        if not numero:
            errores.append("El número de orden de compra es obligatorio.")

        existente = OrdenCompra.query.filter(OrdenCompra.numero == numero, OrdenCompra.id != oc_id).first()
        if existente:
            errores.append(f"La OC '{numero}' ya existe.")

        if errores:
            for e in errores:
                flash(e, "danger")
            return render_template("oc_form.html", oc=oc, valores=request.form)

        old_numero = oc.numero
        old_proveedor = oc.proveedor
        old_estado = oc.estado

        oc.numero = numero
        oc.proveedor = proveedor
        oc.estado = estado

        audit_log("ordenes_compra", oc.id, "numero", old_numero, numero, current_user.username)
        audit_log("ordenes_compra", oc.id, "proveedor", old_proveedor, proveedor, current_user.username)
        audit_log("ordenes_compra", oc.id, "estado", old_estado, estado, current_user.username)

        db.session.commit()
        flash(f"OC '{numero}' actualizada correctamente.", "success")
        return redirect(url_for("routes.oc_lista"))

    return render_template("oc_form.html", oc=oc)


@routes_bp.route("/oc/eliminar/<int:oc_id>", methods=["POST"])
@login_required
def oc_eliminar(oc_id):
    """Eliminar OC solo si no tiene entradas vinculadas."""
    oc = db.session.get(OrdenCompra, oc_id)
    if not oc:
        flash("Orden de compra no encontrada.", "danger")
        return redirect(url_for("routes.oc_lista"))

    if oc.entradas and len(oc.entradas) > 0:
        flash(f"No se puede eliminar: la OC '{oc.numero}' tiene {len(oc.entradas)} entrada(s) asociada(s).", "danger")
        return redirect(url_for("routes.oc_lista"))

    db.session.delete(oc)
    db.session.commit()
    flash(f"OC '{oc.numero}' eliminada correctamente.", "success")
    return redirect(url_for("routes.oc_lista"))


@routes_bp.route("/oc/cerrar/<int:oc_id>", methods=["POST"])
@login_required
def oc_cerrar(oc_id):
    """Marcar OC como CERRADA."""
    oc = db.session.get(OrdenCompra, oc_id)
    if not oc:
        flash("Orden de compra no encontrada.", "danger")
        return redirect(url_for("routes.oc_lista"))

    old_estado = oc.estado
    oc.estado = "CERRADA"
    audit_log("ordenes_compra", oc.id, "estado", old_estado, "CERRADA", current_user.username)
    db.session.commit()
    flash(f"OC '{oc.numero}' marcada como CERRADA.", "success")
    return redirect(url_for("routes.oc_lista"))


# ===========================================================================
# Administración de Usuarios (solo administrador)
# ===========================================================================

def _admin_required(redirect_to="routes.dashboard"):
    """Decorador que verifica que el usuario actual sea el administrador."""
    if not current_user.is_authenticated:
        return redirect(url_for("routes.login"))
    if not current_user.is_admin:
        flash("Acceso denegado. Solo el administrador puede gestionar usuarios.", "danger")
        return redirect(url_for(redirect_to))
    return None


@routes_bp.route("/admin/usuarios")
@login_required
def admin_usuarios():
    """Lista de usuarios del sistema con paginación."""
    check = _admin_required()
    if check:
        return check
    page = request.args.get("page", 1, type=int)
    per_page = 20
    pagination = User.query.order_by(User.username.asc()).paginate(page=page, per_page=per_page, error_out=False)
    return render_template("admin_usuarios.html", pagination=pagination)


@routes_bp.route("/admin/usuarios/nuevo", methods=["GET", "POST"])
@login_required
def admin_usuario_nuevo():
    """Crear nuevo usuario."""
    check = _admin_required()
    if check:
        return check
    if request.method == "POST":
        username = request.form.get("username", "").strip().lower()
        password = request.form.get("password", "")
        password2 = request.form.get("password2", "")

        errores = []
        if not username or len(username) < 3:
            errores.append("El usuario debe tener al menos 3 caracteres.")
        if User.query.filter_by(username=username).first():
            errores.append(f"El usuario '{username}' ya existe.")
        if len(password) < 4:
            errores.append("La contraseña debe tener al menos 4 caracteres.")
        if password != password2:
            errores.append("Las contraseñas no coinciden.")

        if errores:
            for e in errores:
                flash(e, "danger")
            return render_template("admin_usuario_form.html", valores=request.form)

        user = User(username=username)
        user.set_password(password)
        db.session.add(user)
        db.session.commit()
        flash(f"Usuario '{username}' creado correctamente.", "success")
        return redirect(url_for("routes.admin_usuarios"))

    return render_template("admin_usuario_form.html")


@routes_bp.route("/admin/usuarios/editar/<int:user_id>", methods=["GET", "POST"])
@login_required
def admin_usuario_editar(user_id):
    """Editar usuario existente."""
    check = _admin_required()
    if check:
        return check
    user = db.session.get(User, user_id)
    if not user:
        flash("Usuario no encontrado.", "danger")
        return redirect(url_for("routes.admin_usuarios"))

    if request.method == "POST":
        password = request.form.get("password", "")
        password2 = request.form.get("password2", "")

        errores = []
        if password:
            if len(password) < 4:
                errores.append("La contraseña debe tener al menos 4 caracteres.")
            if password != password2:
                errores.append("Las contraseñas no coinciden.")

        if errores:
            for e in errores:
                flash(e, "danger")
            return render_template("admin_usuario_form.html", usuario=user, valores=request.form)

        if password:
            user.set_password(password)
            db.session.commit()
            flash(f"Contraseña de '{user.username}' actualizada.", "success")
        else:
            flash("No se realizaron cambios (dejar contraseña en blanco).", "info")

        return redirect(url_for("routes.admin_usuarios"))

    return render_template("admin_usuario_form.html", usuario=user)


@routes_bp.route("/admin/usuarios/eliminar/<int:user_id>", methods=["POST"])
@login_required
def admin_usuario_eliminar(user_id):
    """Eliminar usuario."""
    check = _admin_required()
    if check:
        return check
    user = db.session.get(User, user_id)
    if not user:
        flash("Usuario no encontrado.", "danger")
        return redirect(url_for("routes.admin_usuarios"))

    if user.username == User.ADMIN_USERNAME:
        flash("No se puede eliminar al usuario administrador principal.", "danger")
        return redirect(url_for("routes.admin_usuarios"))

    if user.id == current_user.id:
        flash("No puedes eliminar tu propio usuario.", "danger")
        return redirect(url_for("routes.admin_usuarios"))

    db.session.delete(user)
    db.session.commit()
    flash(f"Usuario '{user.username}' eliminado.", "success")
    return redirect(url_for("routes.admin_usuarios"))


# ---------------------------------------------------------------------------
# Auditoría (solo administrador)
# ---------------------------------------------------------------------------

@routes_bp.route("/auditoria")
@login_required
def auditoria():
    """Ver el registro de auditoría de cambios."""
    check = _admin_required()
    if check:
        return check

    tabla = request.args.get("tabla", "").strip()
    usuario = request.args.get("usuario", "").strip()
    fecha_desde = request.args.get("fecha_desde", "").strip()
    fecha_hasta = request.args.get("fecha_hasta", "").strip()
    page = request.args.get("page", 1, type=int)
    per_page = 50

    query = AuditLog.query

    if tabla:
        query = query.filter(AuditLog.tabla == tabla)
    if usuario:
        query = query.filter(AuditLog.usuario.ilike(f"%{usuario}%"))
    if fecha_desde:
        try:
            fd = datetime.strptime(fecha_desde, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            query = query.filter(AuditLog.timestamp >= fd)
        except ValueError:
            pass
    if fecha_hasta:
        try:
            fh = datetime.strptime(fecha_hasta, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            query = query.filter(AuditLog.timestamp <= fh)
        except ValueError:
            pass

    query = query.order_by(AuditLog.timestamp.desc())
    pagination = query.paginate(page=page, per_page=per_page, error_out=False)

    return render_template("auditoria.html", pagination=pagination)
